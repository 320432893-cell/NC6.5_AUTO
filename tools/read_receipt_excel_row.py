import argparse
import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from core.utils import load_config  # noqa: E402


DEFAULT_FIELDS = [
    "到款日期",
    "🟪银行来款名",
    "🟪到账金额",
    "🟪原始金额",
    "手续费",
    "币种",
    "银行",
    "客户编码",
    "是否NC已做过",
]


def main():
    parser = argparse.ArgumentParser(description="Read one receipt Excel row.")
    parser.add_argument("row", type=int)
    parser.add_argument("--config", default="config.json")
    parser.add_argument("--all-fields", action="store_true")
    args = parser.parse_args()

    cfg = load_config(args.config)
    excel_cfg = cfg["receipt_entry"]["excel"]
    workbook = openpyxl.load_workbook(excel_cfg["path"], data_only=True, read_only=True)
    try:
        sheet = workbook[excel_cfg["sheet_name"]]
        headers = [
            sheet.cell(excel_cfg.get("header_row", 1), col).value
            for col in range(1, sheet.max_column + 1)
        ]
        data = {
            headers[col - 1]: sheet.cell(args.row, col).value
            for col in range(1, sheet.max_column + 1)
            if headers[col - 1]
        }
    finally:
        workbook.close()

    if not args.all_fields:
        data = {field: data.get(field) for field in DEFAULT_FIELDS}

    print(json.dumps(data, ensure_ascii=False, default=str, indent=2))


if __name__ == "__main__":
    raise SystemExit(main())
