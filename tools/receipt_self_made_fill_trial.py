import argparse
import ctypes
from ctypes import wintypes
from datetime import datetime
from decimal import Decimal
import json
import os
import re
import subprocess
import sys
import time
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from core.jab_operator import JABOperator  # noqa: E402
from core.receipt_config import ReceiptEntryConfig  # noqa: E402
from core.utils import load_config  # noqa: E402
from tools.read_receipt_excel_row import DEFAULT_FIELDS  # noqa: E402
from tools.receipt_body_table_locator import locate_receipt_body_table  # noqa: E402
from tools.receipt_new_probe import (  # noqa: E402
    collect_receipt_new_windows,
    detect_self_made_entry_state,
)
from tools.receipt_keyboard_utils import (  # noqa: E402
    foreground_matches_window,
    get_clipboard_text,
    restore_clipboard_text,
    send_hotkey_ctrl_a,
    send_hotkey_ctrl_v,
    set_clipboard_text,
)
from tools.receipt_table_cell_probe import select_cell  # noqa: E402


CURRENCY_NAMES = {"USD": "美元", "RMB": "人民币", "CNY": "人民币"}
HEADER_DYNAMIC_PREFIX_BASE = "0.0.1.0.0.0.0"
HEADER_COMMON_SUFFIX_TEMPLATE = "0.0.0.1.1.0.0.0.0.1.0.2.0.0.0.0.0.0.0.{index}.0"
HEADER_COMMON_LABEL_SUFFIX_TEMPLATE = "0.0.0.1.1.0.0.0.0.1.0.2.0.0.0.0.0.0.0.{index}"
FINANCE_ORG_LABEL_SUFFIX = "0.0.0.1.1.0.0.0.1.1.1.0"
HEADER_LIVE_SEMANTIC_FALLBACK_TIMEOUT = 0.35
HEADER_FORM_TEXT_INDEXES = {
    "单据日期": 5,
    "币种": 13,
    "收款银行账户": 15,
    "客户": 17,
    "结算方式": 31,
}
HEADER_REQUIRED_LABELS = ("财务组织", "客户", "单据日期", "币种", "结算方式")
HEADER_LABEL_ALIASES = {
    "财务组织": ("财务组织(O)",),
    "客户": ("客户",),
    "单据日期": ("单据日期",),
    "币种": ("币种",),
    "结算方式": ("结算方式",),
}
HEADER_PROBE_LABEL_KEYS = {
    "finance": "财务组织",
    "finance_org": "财务组织",
    "customer": "客户",
    "client": "客户",
    "date": "单据日期",
    "document_date": "单据日期",
    "currency": "币种",
    "settlement": "结算方式",
    "settlement_method": "结算方式",
}
HEADER_SCOPE_ANCHOR_LABEL = "财务组织"
HEADER_SCOPE_ANCHOR_TEXT = "财务组织(O)"
FINANCE_ORG_ACCEPTED_TEXT = "上海移为通信技术股份有限公司"


def print_json(data):
    text = json.dumps(data, ensure_ascii=False, indent=2, default=str)
    try:
        sys.stdout.write(text + "\n")
    except UnicodeEncodeError:
        sys.stdout.buffer.write((text + "\n").encode("utf-8"))


def main():
    parser = argparse.ArgumentParser(
        description="Trial-fill NC receipt self-made entry from one Excel row."
    )
    parser.add_argument("row", type=int, nargs="?")
    parser.add_argument("--config", default="config.json")
    parser.add_argument("--open-self-made", action="store_true")
    parser.add_argument(
        "--probe-header-path-transitions",
        action="store_true",
        help=(
            "只读探测表头 path：当前状态一次，回车后一次，再回车后一次；"
            "用于人工在两次回车之间写财务组织、加窗口"
        ),
    )
    parser.add_argument(
        "--probe-header-semantic-field",
        default=None,
        help="只读测速单个表头字段语义定位，例如 客户、单据日期、币种、结算方式。",
    )
    parser.add_argument(
        "--probe-customer-name-readback",
        action="store_true",
        help="只读探测客户回车后的 NC 客户名称候选来源；不输入、不保存。",
    )
    parser.add_argument(
        "--probe-timeout",
        type=float,
        default=HEADER_LIVE_SEMANTIC_FALLBACK_TIMEOUT,
        help="单次语义定位超时秒数，默认使用正式兜底超时。",
    )
    parser.add_argument(
        "--probe-repeat",
        type=int,
        default=1,
        help="重复测速次数，默认 1。",
    )
    parser.add_argument(
        "--fill-detail",
        action="store_true",
        help="fill receipt detail cells after header verification",
    )
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args()

    config = load_config(args.config)
    if args.probe_header_path_transitions:
        report = probe_header_path_transitions(config)
        print_json(report)
        return 0 if report.get("ok") else 1
    if args.probe_header_semantic_field:
        report = probe_header_semantic_field_speed(
            config,
            normalize_header_probe_label(args.probe_header_semantic_field),
            timeout=args.probe_timeout,
            repeat=args.probe_repeat,
        )
        print_json(report)
        return 0 if report.get("ok") else 1
    if args.probe_customer_name_readback:
        report = probe_customer_name_readback(config, timeout=args.probe_timeout)
        print_json(report)
        return 0 if report.get("ok") else 1
    if args.row is None:
        parser.error(
            "row is required unless --probe-header-path-transitions "
            "or --probe-header-semantic-field or --probe-customer-name-readback is used"
        )
    row_data = read_excel_row(config, args.row)
    business = build_business_values(config, row_data)
    report = {"row": args.row, "excel": row_data, "business": business, "steps": []}

    if args.open_self_made:
        open_step = detect_existing_self_made_entry(config)
        if not open_step.get("ok"):
            open_step = run_receipt_new_probe()
        report["steps"].append(open_step)
        if not open_step.get("ok"):
            report["stopped"] = "open_self_made"
            print_json(report)
            return 1

    jab = JABOperator(config)
    try:
        jab.ensure_started()
        header_steps = fill_header(jab, business)
        report["steps"].extend(header_steps)
        if any(step.get("step") == "blocked" for step in header_steps):
            report["stopped"] = "header"
            print_json(report)
            return 1
        report["steps"].append(read_body_table(jab, "before_detail_fill"))
        if args.fill_detail:
            report["steps"].extend(fill_detail_line(jab, business))
            report["steps"].append(read_body_table(jab, "after_detail_fill"))
        else:
            report["steps"].append(
                {
                    "step": "detail_fill",
                    "ok": False,
                    "blocked": True,
                    "reason": "detail fill requires explicit --fill-detail",
                }
            )
    finally:
        jab.close()

    print_json(report)
    return 0


def read_excel_row(config, row):
    import openpyxl

    excel_cfg = config["receipt_entry"]["excel"]
    workbook = openpyxl.load_workbook(excel_cfg["path"], data_only=True, read_only=True)
    try:
        sheet = workbook[excel_cfg["sheet_name"]]
        header_row = excel_cfg.get("header_row", 1)
        headers = [
            sheet.cell(header_row, col).value for col in range(1, sheet.max_column + 1)
        ]
        data = {
            headers[col - 1]: sheet.cell(row, col).value
            for col in range(1, sheet.max_column + 1)
            if headers[col - 1]
        }
    finally:
        workbook.close()
    return {field: data.get(field) for field in DEFAULT_FIELDS}


def build_business_values(config, row_data):
    receipt_config = ReceiptEntryConfig(config)
    bank = str(row_data.get("银行") or "").strip()
    account = receipt_config.accounts_by_label.get(
        bank.upper()
    ) or receipt_config.accounts_by_label.get(bank)
    if not account:
        normalized_bank = "".join(
            ch for ch in bank.upper() if ch.isalnum() or "\u4e00" <= ch <= "\u9fff"
        )
        account = receipt_config.accounts_by_label.get(normalized_bank)
    if not account:
        raise SystemExit(f"bank account config not found: {bank!r}")
    organization = receipt_config.organizations[account.organization_code]
    currency_code = str(row_data.get("币种") or "").strip().upper()
    receipt_date = row_data.get("到款日期")
    if isinstance(receipt_date, datetime):
        receipt_date_text = receipt_date.strftime("%Y-%m-%d")
    else:
        receipt_date_text = str(receipt_date)[:10]
    amount = Decimal(str(row_data.get("🟪到账金额") or "0"))
    fee_raw = row_data.get("手续费")
    fee = Decimal(str(fee_raw or "0"))
    return {
        "finance_org_code": organization.code,
        "finance_org_name": organization.name,
        "document_date": receipt_date_text,
        "customer_code": str(row_data.get("客户编码") or "").strip(),
        "currency": CURRENCY_NAMES.get(currency_code, currency_code),
        "header_currency_code": account.header_currency_code,
        "bank_label": bank,
        "bank_account": account.account_no,
        "amount": str(amount),
        "fee": str(fee),
        "has_fee": fee != 0,
        "settlement": "网银",
        "main_subject": "1002",
        "main_business_type": "货款",
        "fee_subject": "660305",
        "fee_business_type": "手续费",
    }


def normalize_header_probe_label(label):
    text = str(label or "").strip()
    return HEADER_PROBE_LABEL_KEYS.get(text.lower(), text)


def run_receipt_new_probe():
    return run_receipt_new_probe_with_jab()


def run_receipt_new_probe_with_jab(jab=None):
    if jab is not None:
        from argparse import Namespace
        from tools import receipt_new_probe

        args = Namespace(
            config="config.json",
            method="button",
            path=None,
            title=None,
            class_name="SunAwtFrame",
            name="新增",
            role=None,
            action=None,
            return_timeout=0.2,
            wait=0.8,
            choose_self_made=True,
            self_made_index=0,
            json=False,
            summary=True,
        )
        report = receipt_new_probe.run(args, jab=jab)
        entry_state = report.get("entry_state") or {}
        ok = bool(
            (report.get("open") or {}).get("ok")
            and (report.get("choose_self_made") or {}).get("ok")
        )
        return {
            "step": "open_self_made",
            "ok": ok,
            "parsed": report,
            "entry_state": entry_state,
            "windows_after_choose": report.get("windows_after_choose"),
            "windows_after_open": report.get("windows_after_open"),
            "reason": None
            if ok
            else (
                ((report.get("open") or {}).get("reason"))
                or ((report.get("choose_self_made") or {}).get("reason"))
                or "未能完成新增->自制"
            ),
        }

    cmd = [
        sys.executable,
        str(ROOT / "tools" / "receipt_new_probe.py"),
        "--method",
        "button",
        "--class-name",
        "SunAwtFrame",
        "--choose-self-made",
        "--wait",
        "0.8",
        "--summary",
    ]
    proc = subprocess.run(
        cmd,
        text=True,
        encoding="utf-8",
        errors="replace",
        capture_output=True,
    )
    parsed = None
    ok = False
    try:
        parsed = json.loads(proc.stdout)
        entry_state = parsed.get("entry_state") or {}
        ok = (
            proc.returncode == 0
            and bool((parsed.get("open") or {}).get("ok"))
            and bool((parsed.get("choose_self_made") or {}).get("ok"))
        )
    except json.JSONDecodeError:
        ok = False
    return {
        "step": "open_self_made",
        "ok": ok,
        "returncode": proc.returncode,
        "parsed": parsed,
        "stdout": proc.stdout,
        "stderr": proc.stderr,
    }


def detect_existing_self_made_entry(config):
    jab = JABOperator(config)
    jab.hide_blank_awt_windows_enabled = False
    try:
        jab.ensure_started()
        windows = collect_receipt_new_windows(jab)
        entry_state = detect_self_made_entry_state(windows)
        java_windows = [window for window in windows if window.get("is_java")]
        return {
            "step": "open_self_made",
            "ok": bool(entry_state.get("ok")),
            "method": "existing_entry_state_probe",
            "entry_state": entry_state,
            "java_window_count": len(java_windows),
            "reason": (
                "already in self-made entry state"
                if entry_state.get("ok")
                else "self-made entry state not detected before opening"
            ),
        }
    finally:
        jab.hide_blank_awt_windows_enabled = False
        jab.close()


def probe_header_path_transitions(config):
    jab = JABOperator(config)
    jab.hide_blank_awt_windows_enabled = False
    stages = []
    try:
        jab.ensure_started()
        stages.append(probe_current_header_paths(jab, "before_change"))
        input("已完成第 1 次只读探测。请人工写入/确认财务组织后按回车继续第 2 次探测: ")
        stages.append(probe_current_header_paths(jab, "after_finance"))
        input("已完成第 2 次只读探测。请人工新增/切换窗口后按回车继续第 3 次探测: ")
        stages.append(probe_current_header_paths(jab, "after_window_change"))
        return {
            "ok": any(stage.get("ok") for stage in stages),
            "mode": "header-path-transition-probe",
            "readonly": True,
            "stages": stages,
        }
    finally:
        jab.hide_blank_awt_windows_enabled = False
        jab.close()


def probe_header_semantic_field_speed(config, label, timeout=None, repeat=1):
    jab = JABOperator(config)
    jab.hide_blank_awt_windows_enabled = False
    attempts = []
    timeout = HEADER_LIVE_SEMANTIC_FALLBACK_TIMEOUT if timeout is None else timeout
    repeat = max(int(repeat or 1), 1)
    try:
        jab.ensure_started()
        for index in range(repeat):
            started_at = time.perf_counter()
            found = find_receipt_header_field_by_semantic_label(
                jab,
                label,
                timeout=timeout,
            )
            seconds = round(time.perf_counter() - started_at, 3)
            attempt = {
                "ok": bool(found.get("ok")),
                "index": index + 1,
                "label": label,
                "timeout": timeout,
                "seconds": seconds,
                "source": found.get("source") or "semantic-label-path",
                "path": found.get("path"),
                "label_path": found.get("label_path"),
                "window": found.get("window"),
                "reason": found.get("reason"),
            }
            attempts.append(attempt)
            if found.get("ok"):
                jab.release_contexts(found["vm_id"], found["owned_contexts"])
        successful = [item for item in attempts if item.get("ok")]
        return {
            "ok": bool(successful),
            "mode": "header-semantic-field-speed",
            "readonly": True,
            "label": label,
            "timeout": timeout,
            "repeat": repeat,
            "success_count": len(successful),
            "min_seconds": min((item["seconds"] for item in successful), default=None),
            "max_seconds": max((item["seconds"] for item in successful), default=None),
            "attempts": attempts,
        }
    finally:
        jab.hide_blank_awt_windows_enabled = False
        jab.close()


def probe_customer_name_readback(config, timeout=None):
    jab = JABOperator(config)
    jab.hide_blank_awt_windows_enabled = False
    timeout = HEADER_LIVE_SEMANTIC_FALLBACK_TIMEOUT if timeout is None else timeout
    try:
        jab.ensure_started()
        scope = resolve_current_header_scope_for_probe(jab, timeout=timeout)
        candidates = []
        if scope.get("ok"):
            candidates.extend(
                collect_customer_field_candidates_for_scope(
                    jab,
                    scope.get("scope_hwnd"),
                    scope.get("dynamic_index"),
                    timeout=timeout,
                )
            )
        best = first_valid_customer_name(candidates)
        return {
            "ok": bool(best),
            "mode": "customer-name-readback-probe",
            "readonly": True,
            "scope": scope,
            "best": best,
            "field_candidates": candidates,
            "reason": None if best else "未读到有效 NC 客户名称候选",
        }
    finally:
        jab.hide_blank_awt_windows_enabled = False
        jab.close()


def resolve_current_header_scope_for_probe(jab, timeout=None):
    windows = collect_receipt_new_windows(jab)
    state = detect_self_made_entry_state(windows)
    scope_hwnd = None
    for hit in state.get("hits") or []:
        window = hit.get("window") or {}
        if window.get("class_name") == "SunAwtCanvas" and window.get("hwnd"):
            scope_hwnd = int(window["hwnd"])
            break
    if scope_hwnd:
        anchor = resolve_receipt_header_anchor_in_canvas(
            jab,
            scope_hwnd,
            timeout=timeout or 0.6,
        )
        if anchor.get("ok"):
            return anchor
    semantic = infer_receipt_header_scope_by_semantic(jab, scope_hwnd=scope_hwnd)
    if semantic.get("ok"):
        semantic["entry_state"] = state
        return semantic
    return {
        "ok": False,
        "reason": "当前页面未解析到收款单表头 scope",
        "scope_hwnd": scope_hwnd,
        "entry_state": state,
        "semantic_attempt": semantic,
    }


def collect_customer_field_candidates_for_scope(
    jab, scope_hwnd, dynamic_index, timeout
):
    candidates = []
    path_found = find_receipt_header_field_by_dynamic_path(
        jab,
        "客户",
        dynamic_index,
        scope_hwnd=scope_hwnd,
        require_showing=False,
        require_valid_bounds=False,
    )
    candidates.extend(snapshot_header_field_candidate(jab, path_found, "path"))
    semantic_found = find_receipt_header_field_by_semantic_label(
        jab,
        "客户",
        scope_hwnd=scope_hwnd,
        timeout=timeout,
    )
    candidates.extend(
        snapshot_header_field_candidate(jab, semantic_found, "semantic-label-path")
    )
    return dedupe_customer_candidates(candidates)


def snapshot_header_field_candidate(jab, found, source):
    if not found.get("ok"):
        return [
            {
                "ok": False,
                "source": source,
                "reason": found.get("reason"),
                "path": found.get("path"),
                "label_path": found.get("label_path"),
            }
        ]
    context = found["context"]
    vm_id = found["vm_id"]
    owned_contexts = found["owned_contexts"]
    try:
        info = jab.get_context_info(vm_id, context)
        text = jab.get_text_context_value(vm_id, context)
        snapshot = context_text_snapshot(info, text)
        return [
            {
                "ok": True,
                "source": source,
                "path": found.get("path"),
                "label_path": found.get("label_path"),
                "window": found.get("window"),
                "role": snapshot.get("role"),
                "states": snapshot.get("states"),
                "text": snapshot.get("text"),
                "name": snapshot.get("name"),
                "description": snapshot.get("description"),
                "valid_values": valid_customer_values_from_snapshot(snapshot),
            }
        ]
    finally:
        jab.release_contexts(vm_id, owned_contexts)


def collect_customer_nearby_candidates(jab, label_path, scope_hwnd, source):
    if not label_path:
        return []
    parent_path = ".".join(str(label_path).split(".")[:-1])
    if not parent_path:
        return []
    context, vm_id, owned_contexts, window_info = jab.find_context_by_path_once(
        parent_path,
        class_name="SunAwtCanvas",
        scope_hwnd=scope_hwnd,
        require_showing=False,
        require_valid_bounds=False,
    )
    if not context:
        return [
            {
                "ok": False,
                "source": source,
                "reason": "label parent path not found",
                "parent_path": parent_path,
                "label_path": label_path,
            }
        ]
    try:
        rows = []
        collect_context_text_rows(
            jab,
            vm_id,
            context,
            parent_path,
            depth=0,
            max_depth=4,
            rows=rows,
        )
        candidates = []
        for row in rows:
            values = valid_customer_values_from_snapshot(row)
            if values or row.get("path") == label_path:
                candidates.append(
                    {
                        "ok": True,
                        "source": source,
                        "parent_path": parent_path,
                        "label_path": label_path,
                        "window": window_info,
                        **row,
                        "valid_values": values,
                    }
                )
        return candidates
    finally:
        jab.release_contexts(vm_id, owned_contexts)


def collect_context_text_rows(jab, vm_id, context, path, depth, max_depth, rows):
    info = jab.get_context_info(vm_id, context)
    if not info:
        return
    text = jab.get_text_context_value(vm_id, context)
    snapshot = context_text_snapshot(info, text)
    if any(snapshot.get(key) for key in ("text", "name", "description")):
        rows.append({"path": path, "depth": depth, **snapshot})
    if depth >= max_depth:
        return
    role = (info.role_en_US.strip() or info.role.strip()).lower()
    if role == "table":
        return
    for index in range(min(info.childrenCount, jab.max_children)):
        child = jab.dll.getAccessibleChildFromContext(vm_id, context, index)
        if not child:
            continue
        try:
            collect_context_text_rows(
                jab,
                vm_id,
                child,
                f"{path}.{index}",
                depth + 1,
                max_depth,
                rows,
            )
        finally:
            jab.release_contexts(vm_id, [child])


def context_text_snapshot(info, text):
    return {
        "role": (info.role_en_US.strip() or info.role.strip()) if info else "",
        "states": (info.states_en_US.strip() or info.states.strip()) if info else "",
        "text": str(text or "").strip(),
        "name": info.name.strip() if info else "",
        "description": info.description.strip() if info else "",
    }


def valid_customer_values_from_snapshot(snapshot):
    values = []
    for key in ("description", "text", "name"):
        value = str((snapshot or {}).get(key) or "").strip()
        if is_valid_customer_name_candidate(value) and value not in values:
            values.append(value)
    return values


def is_valid_customer_name_candidate(value):
    text = str(value or "").strip()
    if not text:
        return False
    if re.match(r"^\[L?java(\.|x\.)", text) or re.match(
        r"^\[L[^;]+;@[0-9a-fA-F]+$", text
    ):
        return False
    if re.match(r"^[A-Z]{1,5}\d{3,}$", text):
        return False
    if text in {"客户", "客户编码"}:
        return False
    return bool(re.search(r"[\u4e00-\u9fffA-Za-z]", text))


def first_valid_customer_name(candidates):
    for candidate in candidates:
        values = candidate.get("valid_values") or []
        if values:
            return {
                "value": values[0],
                "source": candidate.get("source"),
                "path": candidate.get("path"),
                "parent_path": candidate.get("parent_path"),
                "field": next(
                    (
                        key
                        for key in ("description", "text", "name")
                        if candidate.get(key) == values[0]
                    ),
                    None,
                ),
            }
    return None


def dedupe_customer_candidates(candidates):
    seen = set()
    unique = []
    for item in candidates:
        key = (
            item.get("source"),
            item.get("path"),
            item.get("parent_path"),
            item.get("text"),
            item.get("name"),
            item.get("description"),
            item.get("reason"),
        )
        if key in seen:
            continue
        seen.add(key)
        unique.append(item)
    return unique


def probe_current_header_paths(jab, stage):
    started_at = time.perf_counter()
    windows = collect_receipt_new_windows(jab)
    candidates = []
    for window in windows:
        if not window.get("visible") or window.get("class_name") != "SunAwtCanvas":
            continue
        hwnd = window.get("hwnd")
        if not hwnd:
            continue
        anchor = resolve_receipt_header_anchor_in_canvas(jab, int(hwnd), timeout=0.15)
        if not anchor.get("ok"):
            continue
        dynamic_index = anchor.get("dynamic_index")
        samples = probe_header_field_paths_for_scope(
            jab,
            int(hwnd),
            dynamic_index,
        )
        candidates.append(
            {
                "scope_hwnd": int(hwnd),
                "dynamic_index": dynamic_index,
                "dynamic_prefix": receipt_header_dynamic_prefix(dynamic_index),
                "anchor": anchor,
                "fields": samples,
            }
        )
    return {
        "ok": bool(candidates),
        "stage": stage,
        "seconds": round(time.perf_counter() - started_at, 3),
        "candidate_count": len(candidates),
        "candidates": candidates,
    }


def probe_header_field_paths_for_scope(jab, scope_hwnd, dynamic_index):
    fields = {}
    for label in HEADER_REQUIRED_LABELS:
        fields[label] = {
            "default_path": summarize_header_path_attempt(
                jab,
                label,
                dynamic_index,
                scope_hwnd,
                template=receipt_header_default_path_template(),
            ),
            "semantic_path": summarize_semantic_header_attempt(
                jab,
                label,
                scope_hwnd,
            ),
        }
    return fields


def summarize_header_path_attempt(jab, label, dynamic_index, scope_hwnd, template):
    if label == HEADER_SCOPE_ANCHOR_LABEL:
        path = build_receipt_header_dynamic_path(dynamic_index, label)
        label_path = build_receipt_header_dynamic_label_path(dynamic_index, label)
        template_source = "finance-anchor"
    else:
        path = build_receipt_header_path_from_template(
            dynamic_index,
            label,
            template,
        )
        label_path = build_receipt_header_label_path_from_template(
            dynamic_index,
            label,
            template,
        )
        template_source = (template or {}).get("source")
    if not path:
        return {
            "ok": False,
            "label": label,
            "reason": "path not configured",
            "template_source": template_source,
        }
    found = find_context_by_path_readonly(
        jab,
        path,
        scope_hwnd=scope_hwnd,
        role="text",
    )
    return {
        "label": label,
        "ok": bool(found.get("ok")),
        "path": path,
        "label_path": label_path,
        "template_source": template_source,
        **found,
    }


def summarize_semantic_header_attempt(jab, label, scope_hwnd):
    found = find_receipt_header_field_by_semantic_label(
        jab,
        label,
        scope_hwnd=scope_hwnd,
    )
    summary = {
        "label": label,
        "ok": bool(found.get("ok")),
        "source": found.get("source") or "semantic-label-path",
        "path": found.get("path"),
        "label_path": found.get("label_path"),
        "window": found.get("window"),
        "reason": found.get("reason"),
    }
    if found.get("ok"):
        inferred = infer_header_path_template_from_field(
            found.get("path"),
            extract_receipt_header_dynamic_index(found.get("path")),
            label,
        )
        summary["inferred_template"] = inferred
        jab.release_contexts(found["vm_id"], found["owned_contexts"])
    return summary


def find_context_by_path_readonly(jab, path, scope_hwnd=None, role=None):
    context, vm_id, owned, window_info = jab.find_context_by_path_once(
        path,
        class_name="SunAwtCanvas",
        scope_hwnd=scope_hwnd,
        role=role,
        require_showing=False,
        require_valid_bounds=False,
    )
    if not context:
        return {"ok": False, "reason": "path not found"}
    try:
        info = jab.get_context_info(vm_id, context)
        text = jab.get_text_context_value(vm_id, context)
        return {
            "ok": True,
            "window": window_info,
            "name": info.name.strip() if info else "",
            "description": info.description.strip() if info else "",
            "role": (info.role_en_US.strip() or info.role.strip()) if info else "",
            "states": (
                (info.states_en_US.strip() or info.states.strip()) if info else ""
            ),
            "text": text,
        }
    finally:
        jab.release_contexts(vm_id, owned)


def fill_header(
    jab,
    business,
    after_field=None,
    scope_hwnd=None,
    dynamic_index=None,
    anchor_path=None,
    recover_after_failure=None,
    trust_provided_scope=False,
):
    started_at = time.perf_counter()
    steps = []
    scope_started_at = time.perf_counter()
    scope = resolve_receipt_header_scope(
        jab,
        scope_hwnd,
        dynamic_index,
        anchor_path,
        trust_provided_scope=trust_provided_scope,
    )
    scope_seconds = round(time.perf_counter() - scope_started_at, 3)
    if not scope.get("ok"):
        return [
            {
                "step": "blocked",
                "reason": "header scope not resolved",
                "scope": scope,
                "header_timing": {
                    "scope_seconds": scope_seconds,
                    "total_seconds": round(time.perf_counter() - started_at, 3),
                },
            }
        ]
    dynamic_index = scope.get("dynamic_index")
    scope_hwnd = scope.get("scope_hwnd")
    learned_header_template = None
    if dynamic_index is not None:
        clear_receipt_header_path_template_cache()
    for field in [
        {
            "label": "财务组织",
            "value": business["finance_org_code"],
            "accepted_text": business.get("finance_org_name"),
            "dynamic_path": True,
        },
        {
            "label": "客户",
            "value": business["customer_code"],
            "dynamic_path": True,
        },
        {
            "label": "单据日期",
            "value": business["document_date"],
            "dynamic_path": True,
        },
        {
            "label": "币种",
            "value": business.get("currency"),
            "accepted_text": business.get("currency"),
            "dynamic_path": True,
        },
        {
            "label": "结算方式",
            "value": business.get("settlement") or "网银",
            "dynamic_path": True,
        },
    ]:
        label = field["label"]
        value = field["value"]
        if field.get("dynamic_path"):
            field_started_at = time.perf_counter()
            result = set_receipt_header_dynamic_field(
                jab,
                label,
                value,
                dynamic_index,
                scope_hwnd,
                accepted_text=field.get("accepted_text"),
                recover_after_failure=recover_after_failure,
                path_template=learned_header_template,
            )
            ok = bool(result.get("ok"))
            if ok and label != HEADER_SCOPE_ANCHOR_LABEL:
                inferred_template = infer_header_path_template_from_field(
                    result.get("path"),
                    dynamic_index,
                    label,
                )
                if inferred_template and learned_header_template is None:
                    learned_header_template = inferred_template
                    set_receipt_header_path_template(dynamic_index, inferred_template)
                    result["header_path_template_learned"] = learned_header_template
            field_seconds = round(time.perf_counter() - field_started_at, 3)
            steps.append(
                {
                    "step": "header",
                    "label": label,
                    "value": value,
                    "method": "dynamic_path_commit",
                    "scope": {
                        "mode": scope.get("mode"),
                        "scope_hwnd": scope_hwnd,
                        "dynamic_index": dynamic_index,
                        "dynamic_prefix": scope.get("dynamic_prefix"),
                    },
                    **result,
                    "seconds": field_seconds,
                    "header_timing": {
                        "scope_seconds": scope_seconds,
                        "elapsed_seconds": round(time.perf_counter() - started_at, 3),
                    },
                }
            )
        else:
            ok = False
            steps.append(
                {
                    "step": "header",
                    "label": label,
                    "value": value,
                    "method": "unsupported",
                    "ok": False,
                    "reason": "header field must use dynamic path",
                }
            )
        if not ok:
            steps.append(
                {
                    "step": "blocked",
                    "reason": "header field failed; blocking workflow stopped",
                    "label": label,
                }
            )
            break
        if after_field is not None:
            callback_report = after_field(label, value, steps[-1])
            if callback_report is not None:
                steps[-1]["after_field_callback"] = callback_report
                if not callback_report.get("ok", True):
                    steps.append(
                        {
                            "step": "blocked",
                            "reason": callback_report.get("reason")
                            or "header after-field callback blocked workflow",
                            "label": label,
                        }
                    )
                    break
    return steps


def resolve_receipt_header_scope(
    jab,
    scope_hwnd=None,
    dynamic_index=None,
    anchor_path=None,
    trust_provided_scope=False,
):
    cached = getattr(jab, "_receipt_header_scope_cache", None)
    if cached and (scope_hwnd is None or cached.get("scope_hwnd") == scope_hwnd):
        return {**cached, "cached": True}
    if scope_hwnd and dynamic_index is not None:
        if trust_provided_scope:
            scoped = {
                "ok": True,
                "scope_hwnd": scope_hwnd,
                "mode": "provided-canvas-anchor-trusted",
                "dynamic_index": dynamic_index,
                "dynamic_prefix": receipt_header_dynamic_prefix(dynamic_index),
                "matched_labels": [HEADER_SCOPE_ANCHOR_LABEL],
                "semantic_label_path": anchor_path,
                "label_path": anchor_path,
            }
            try:
                setattr(jab, "_receipt_header_scope_cache", scoped)
            except AttributeError:
                pass
            return scoped
        scoped = validate_receipt_header_scope_anchor(
            jab,
            scope_hwnd,
            dynamic_index,
            anchor_path=anchor_path,
        )
        if scoped.get("ok"):
            try:
                setattr(jab, "_receipt_header_scope_cache", scoped)
            except AttributeError:
                pass
            return scoped
        return scoped
    return {
        "ok": False,
        "mode": "provided-canvas-anchor",
        "reason": "正式表头缺少当前 canvas scope 或 dynamic_index，停止；不走语义兜底",
        "scope_hwnd": scope_hwnd,
        "dynamic_index": dynamic_index,
        "label_path": anchor_path,
    }


def validate_receipt_header_scope_anchor(
    jab, scope_hwnd, dynamic_index, anchor_path=None
):
    label_path = anchor_path or build_receipt_header_dynamic_label_path(
        dynamic_index,
        HEADER_SCOPE_ANCHOR_LABEL,
    )
    if not label_path:
        return {
            "ok": False,
            "mode": "provided-canvas-anchor",
            "reason": "财务组织(O) label path not configured",
            "scope_hwnd": scope_hwnd,
            "dynamic_index": dynamic_index,
        }
    context, vm_id, owned_contexts, window_info = jab.find_context_by_path_once(
        label_path,
        class_name="SunAwtCanvas",
        scope_hwnd=scope_hwnd,
        role="label",
        require_showing=False,
        require_valid_bounds=False,
    )
    if not context:
        return {
            "ok": False,
            "mode": "provided-canvas-anchor",
            "reason": "当前 canvas 未找到财务组织(O) 锚点",
            "scope_hwnd": scope_hwnd,
            "dynamic_index": dynamic_index,
            "label_path": label_path,
        }
    try:
        info = jab.get_context_info(vm_id, context)
        anchor_ok = bool(info and header_scope_anchor_text_matches(info))
        anchor_text = {
            "name": info.name.strip() if info else "",
            "description": info.description.strip() if info else "",
        }
    finally:
        jab.release_contexts(vm_id, owned_contexts)
    if not anchor_ok:
        return {
            "ok": False,
            "mode": "provided-canvas-anchor",
            "reason": "当前 canvas 财务组织(O) 锚点文本不匹配",
            "scope_hwnd": scope_hwnd,
            "dynamic_index": dynamic_index,
            "label_path": label_path,
            "anchor_text": anchor_text,
        }
    return {
        "ok": True,
        "scope_hwnd": scope_hwnd,
        "mode": "provided-canvas-anchor",
        "dynamic_index": dynamic_index,
        "dynamic_prefix": receipt_header_dynamic_prefix(dynamic_index),
        "matched_labels": [HEADER_SCOPE_ANCHOR_LABEL],
        "semantic_label_path": label_path,
        "anchor_text": anchor_text,
        "window": window_info,
    }


def resolve_receipt_header_anchor_in_canvas(jab, scope_hwnd, timeout=0.6):
    label_found = find_header_label_context_with_window(
        jab,
        HEADER_SCOPE_ANCHOR_LABEL,
        timeout=timeout,
        require_showing=False,
        scope_hwnd=scope_hwnd,
        strict_anchor=True,
    )
    label_context, vm_id, owned_contexts, owned_indexes, window = label_found
    if not label_context:
        return {
            "ok": False,
            "reason": "当前 canvas 未找到财务组织(O) 锚点",
            "scope_hwnd": scope_hwnd,
            "window": window,
        }
    label_path = ".".join(["0", *[str(index) for index in owned_indexes]])
    try:
        info = jab.get_context_info(vm_id, label_context)
        anchor_ok = bool(info and header_scope_anchor_text_matches(info))
        anchor_text = {
            "name": info.name.strip() if info else "",
            "description": info.description.strip() if info else "",
        }
    finally:
        jab.release_contexts(vm_id, owned_contexts)
    if not anchor_ok:
        return {
            "ok": False,
            "reason": "当前 canvas 财务组织(O) 锚点文本不匹配",
            "scope_hwnd": scope_hwnd,
            "label_path": label_path,
            "anchor_text": anchor_text,
            "window": window,
        }
    dynamic_index = extract_receipt_header_dynamic_index(label_path)
    if dynamic_index is None:
        return {
            "ok": False,
            "reason": "当前 canvas 财务组织(O) 锚点无法推出动态前缀",
            "scope_hwnd": scope_hwnd,
            "label_path": label_path,
            "anchor_text": anchor_text,
            "window": window,
        }
    corrected = correct_header_anchor_dynamic_index_by_customer(
        jab,
        scope_hwnd,
        dynamic_index,
    )
    if corrected.get("ok") and corrected.get("dynamic_index") != dynamic_index:
        corrected_index = corrected.get("dynamic_index")
        return {
            "ok": True,
            "scope_hwnd": scope_hwnd,
            "dynamic_index": corrected_index,
            "dynamic_prefix": receipt_header_dynamic_prefix(corrected_index),
            "label_path": label_path,
            "anchor_text": anchor_text,
            "window": corrected.get("window") or window,
            "mode": "current-canvas-anchor-corrected-by-customer",
            "initial_dynamic_index": dynamic_index,
            "initial_dynamic_prefix": receipt_header_dynamic_prefix(dynamic_index),
            "correction": corrected,
        }
    return {
        "ok": True,
        "scope_hwnd": scope_hwnd,
        "dynamic_index": dynamic_index,
        "dynamic_prefix": receipt_header_dynamic_prefix(dynamic_index),
        "label_path": label_path,
        "anchor_text": anchor_text,
        "window": window,
        "mode": "current-canvas-anchor",
    }


def correct_header_anchor_dynamic_index_by_customer(jab, scope_hwnd, dynamic_index):
    current = find_receipt_header_field_by_dynamic_path(
        jab,
        "客户",
        dynamic_index,
        scope_hwnd=scope_hwnd,
        require_showing=False,
        require_valid_bounds=False,
    )
    if current.get("ok"):
        jab.release_contexts(current["vm_id"], current["owned_contexts"])
        return {
            "ok": True,
            "source": "current-anchor-customer-path",
            "dynamic_index": dynamic_index,
            "path": current.get("path"),
            "window": current.get("window"),
        }
    semantic = find_receipt_header_field_by_semantic_label(
        jab,
        "客户",
        scope_hwnd=scope_hwnd,
    )
    if not semantic.get("ok"):
        return {
            "ok": False,
            "source": "customer-semantic-correction",
            "dynamic_index": dynamic_index,
            "current_attempt": current,
            "semantic_attempt": semantic,
        }
    corrected_index = extract_receipt_header_dynamic_index(semantic.get("path"))
    path = semantic.get("path")
    window = semantic.get("window")
    jab.release_contexts(semantic["vm_id"], semantic["owned_contexts"])
    if corrected_index is None:
        return {
            "ok": False,
            "source": "customer-semantic-correction",
            "dynamic_index": dynamic_index,
            "current_attempt": current,
            "semantic_path": path,
            "reason": "客户语义 path 无法推出 dynamic_index",
        }
    return {
        "ok": True,
        "source": "customer-semantic-correction",
        "dynamic_index": corrected_index,
        "path": path,
        "window": window,
        "current_attempt": current,
    }


def wait_header_account_description(jab, timeout=5.0, scope=None):
    deadline = time.time() + max(float(timeout or 0), 0.0)
    last = None
    if scope is None:
        scope = locate_receipt_header_scope(jab)
    if not scope.get("ok"):
        return {"text": None, "description": "", "accepted": False, "scope": scope}
    dynamic_index = scope.get("dynamic_index")
    scope_hwnd = scope.get("scope_hwnd")
    first_attempt = True
    while first_attempt or time.time() < deadline:
        first_attempt = False
        found = find_receipt_header_field_by_dynamic_path(
            jab,
            "收款银行账户",
            dynamic_index,
            scope_hwnd=scope_hwnd,
            require_showing=False,
            require_valid_bounds=False,
        )
        if found.get("ok"):
            context = found["context"]
            vm_id = found["vm_id"]
            owned_contexts = found["owned_contexts"]
            try:
                info = jab.get_context_info(vm_id, context)
                text = jab.get_text_context_value(vm_id, context)
                desc = info.description.strip() if info else ""
                last = {"text": text, "description": desc, "path": found.get("path")}
                if text or desc:
                    last["accepted"] = True
                    return last
            finally:
                jab.release_contexts(vm_id, owned_contexts)
        if time.time() >= deadline:
            break
        time.sleep(0.3)
    if last is None:
        last = {"text": None, "description": "", "accepted": False}
    else:
        last["accepted"] = False
    return last


def set_receipt_header_dynamic_field(
    jab,
    label,
    value,
    dynamic_index,
    scope_hwnd,
    accepted_text=None,
    recover_after_failure=None,
    path_template=None,
):
    legacy_control_name_attempt = None
    if label == HEADER_SCOPE_ANCHOR_LABEL:
        legacy_control_name_attempt = set_finance_org_by_legacy_control_name(
            jab,
            value,
            scope_hwnd=scope_hwnd,
            accepted_text=accepted_text,
        )
        return {
            **legacy_control_name_attempt,
            "dynamic_index": dynamic_index,
            "dynamic_prefix": receipt_header_dynamic_prefix(dynamic_index),
        }

    def find_by_path():
        path_found = find_receipt_header_field_by_dynamic_path(
            jab,
            label,
            dynamic_index,
            scope_hwnd=scope_hwnd,
            require_showing=False,
            require_valid_bounds=False,
            path_template=path_template,
        )
        if path_found.get("ok"):
            return path_found
        live_found = find_receipt_header_field_by_live_semantic(
            jab,
            label,
            scope_hwnd=scope_hwnd,
            timeout=HEADER_LIVE_SEMANTIC_FALLBACK_TIMEOUT,
            include_scoped=False,
        )
        if live_found.get("ok"):
            live_found["dynamic_path_attempt"] = path_found
            return live_found
        return {
            **live_found,
            "dynamic_path_attempt": path_found,
            "live_semantic_timeout": HEADER_LIVE_SEMANTIC_FALLBACK_TIMEOUT,
        }

    recovery_after_find = None
    try:
        found = find_by_path()
    except Exception as exc:
        if recover_after_failure is None:
            raise
        recovery_after_find = recover_after_failure()
        if recovery_after_find.get("attempted") and recovery_after_find.get("ok"):
            try:
                found = find_by_path()
            except Exception as retry_exc:
                return {
                    "ok": False,
                    "stage": "resolve",
                    "exception": f"{type(retry_exc).__name__}: {retry_exc}",
                    "first_exception": f"{type(exc).__name__}: {exc}",
                    "modal_recovery": recovery_after_find,
                }
        else:
            return {
                "ok": False,
                "stage": "resolve",
                "exception": f"{type(exc).__name__}: {exc}",
                "modal_recovery": recovery_after_find,
            }
    if not found.get("ok") and recover_after_failure is not None:
        recovery_after_find = recover_after_failure()
        if recovery_after_find.get("attempted") and recovery_after_find.get("ok"):
            found = find_by_path()
    path_attempt = found
    if not found.get("ok"):
        return {
            "ok": False,
            "stage": "resolve",
            "path_attempt": path_attempt,
            "modal_recovery": recovery_after_find,
            "legacy_control_name_attempt": legacy_control_name_attempt,
        }
    context = found["context"]
    vm_id = found["vm_id"]
    owned_contexts = found["owned_contexts"]
    window_info = found["window"]

    def write_current_context(initial_recovery=None):
        info_before = jab.get_context_info(vm_id, context)
        before = jab.get_text_context_value(vm_id, context)
        set_text_ok = False
        modal_recovery = initial_recovery
        guarded_paste = guarded_paste_header_value(
            jab,
            vm_id,
            context,
            window_info,
            value,
        )
        set_ok = bool(guarded_paste.get("ok"))
        if not set_ok and recover_after_failure is not None:
            recovery_after_set = recover_after_failure()
            modal_recovery = recovery_after_set or modal_recovery
            if recovery_after_set.get("attempted") and recovery_after_set.get("ok"):
                guarded_paste = guarded_paste_header_value(
                    jab,
                    vm_id,
                    context,
                    window_info,
                    value,
                )
                set_ok = bool(guarded_paste.get("ok"))
        if set_ok and hasattr(jab.dll, "requestFocus"):
            jab.dll.requestFocus(vm_id, context)
        commit_action = None
        enter_ok = (
            bool(guarded_paste.get("enter_ok"))
            if guarded_paste and guarded_paste.get("ok")
            else False
        )
        info_after = jab.get_context_info(vm_id, context)
        after = jab.get_text_context_value(vm_id, context)
        backend_state = describe_backend_field_state(
            info_after,
            after,
            value=value,
            accepted_text=accepted_text,
        )
        acceptance_probe = None
        accepted = bool(backend_state.get("accepted"))
        if set_ok and accepted_text and not accepted:
            acceptance_probe = confirm_header_field_accepted(
                jab,
                vm_id,
                context,
                expected_text=accepted_text,
                value=value,
            )
            accepted = bool(acceptance_probe.get("accepted"))
        return {
            "ok": bool(set_ok and (not accepted_text or accepted)),
            "path": found.get("path"),
            "label_path": found.get("label_path"),
            "dynamic_index": found.get("dynamic_index"),
            "dynamic_prefix": found.get("dynamic_prefix"),
            "source": found.get("source") or "path",
            "path_attempt": path_attempt,
            "dynamic_path_attempt": found.get("dynamic_path_attempt"),
            "text_before": before,
            "description_before": (
                info_before.description.strip() if info_before else None
            ),
            "set_ok": bool(set_ok),
            "set_text_ok": bool(set_text_ok),
            "guarded_paste": guarded_paste,
            "legacy_control_name_attempt": legacy_control_name_attempt,
            "modal_recovery": modal_recovery,
            "commit_action": commit_action,
            "enter_ok": bool(enter_ok),
            "post_write_snapshot": backend_state,
            "acceptance_probe": acceptance_probe,
            "accepted_text": accepted_text_from_backend(
                (acceptance_probe or {}).get("matched_snapshot") or backend_state,
                value,
                accepted_text,
            ),
            "text_after": after,
            "description_after": (
                info_after.description.strip() if info_after else None
            ),
        }

    try:
        try:
            return write_current_context(recovery_after_find)
        except Exception as exc:
            if recover_after_failure is None:
                raise
            recovery_after_exception = recover_after_failure()
            if recovery_after_exception.get(
                "attempted"
            ) and recovery_after_exception.get("ok"):
                try:
                    retried = write_current_context(recovery_after_exception)
                except Exception as retry_exc:
                    return {
                        "ok": False,
                        "stage": "write",
                        "exception": f"{type(retry_exc).__name__}: {retry_exc}",
                        "first_exception": f"{type(exc).__name__}: {exc}",
                        "path": found.get("path"),
                        "label_path": found.get("label_path"),
                        "dynamic_index": found.get("dynamic_index"),
                        "dynamic_prefix": found.get("dynamic_prefix"),
                        "modal_recovery": recovery_after_exception,
                        "legacy_control_name_attempt": legacy_control_name_attempt,
                    }
                retried["retried_after_modal_recovery"] = True
                return retried
            return {
                "ok": False,
                "stage": "write",
                "exception": f"{type(exc).__name__}: {exc}",
                "path": found.get("path"),
                "label_path": found.get("label_path"),
                "dynamic_index": found.get("dynamic_index"),
                "dynamic_prefix": found.get("dynamic_prefix"),
                "modal_recovery": recovery_after_exception,
                "legacy_control_name_attempt": legacy_control_name_attempt,
            }
    finally:
        jab.release_contexts(vm_id, owned_contexts)


def confirm_header_field_accepted(
    jab,
    vm_id,
    context,
    expected_text,
    value=None,
    timeout=1.2,
    interval=0.1,
):
    deadline = time.time() + max(float(timeout or 0), 0.0)
    attempts = []
    first = True
    while first or time.time() < deadline:
        first = False
        info = jab.get_context_info(vm_id, context)
        text = jab.get_text_context_value(vm_id, context)
        snapshot = describe_backend_field_state(
            info,
            text,
            value=value,
            accepted_text=expected_text,
        )
        attempts.append(snapshot)
        if snapshot.get("accepted"):
            return {
                "ok": True,
                "accepted": True,
                "expected_text": expected_text,
                "source": "current-context",
                "attempts": attempts,
                "matched_snapshot": snapshot,
            }
        if time.time() >= deadline:
            break
        time.sleep(min(max(float(interval or 0), 0.02), deadline - time.time()))
    return {
        "ok": False,
        "accepted": False,
        "expected_text": expected_text,
        "attempts": attempts,
        "reason": "表头字段未确认解析为目标值",
    }


def find_receipt_header_field_by_live_semantic(
    jab,
    label,
    scope_hwnd=None,
    timeout=HEADER_LIVE_SEMANTIC_FALLBACK_TIMEOUT,
    include_scoped=False,
):
    found = find_receipt_header_field_by_semantic_label(
        jab,
        label,
        scope_hwnd=scope_hwnd,
        timeout=timeout,
    )
    if found.get("ok"):
        found["source"] = "semantic-live-after-path-miss"
        return found
    if not include_scoped:
        return {
            **found,
            "source": "semantic-live-after-path-miss",
            "timeout": timeout,
        }
    scoped_found = find_receipt_header_field_by_scoped_label(
        jab,
        label,
        scope_hwnd=scope_hwnd,
    )
    if scoped_found.get("ok"):
        scoped_found["source"] = "scoped-label-live-after-path-miss"
        scoped_found["semantic_label_attempt"] = found
        return scoped_found
    return {
        **scoped_found,
        "source": "semantic-live-after-path-miss",
        "semantic_label_attempt": found,
    }


def set_finance_org_by_legacy_control_name(
    jab,
    value,
    scope_hwnd=None,
    accepted_text=None,
):
    accepted_text = accepted_text or FINANCE_ORG_ACCEPTED_TEXT
    context, vm_id, owned_contexts, owned_indexes, window_info = (
        find_context_with_window(
            jab,
            HEADER_SCOPE_ANCHOR_TEXT,
            roles=("text",),
            timeout=1.5,
            require_showing=True,
            window_class="SunAwtCanvas",
            visible_only=True,
            scope_hwnd=scope_hwnd,
        )
    )
    if not context:
        return {
            "ok": False,
            "method": "legacy-control-name-guarded-paste-enter",
            "reason": "control not found",
            "control_name": HEADER_SCOPE_ANCHOR_TEXT,
            "scope_hwnd": scope_hwnd,
        }
    path = "0" + "".join(f".{index}" for index in owned_indexes)
    try:
        info_before = jab.get_context_info(vm_id, context)
        before = jab.get_text_context_value(vm_id, context)
        paste_result = guarded_paste_header_value(
            jab,
            vm_id,
            context,
            window_info,
            value,
        )
        set_text_result = None
        if not paste_result.get("ok"):
            set_ok = bool(jab.set_text_context(vm_id, context, value))
            set_text_result = {"ok": set_ok, "method": "setTextContents"}
        info_after = jab.get_context_info(vm_id, context)
        after = jab.get_text_context_value(vm_id, context)
        backend_state = describe_backend_field_state(
            info_after,
            after,
            value=value,
            accepted_text=accepted_text,
        )
        acceptance_probe = confirm_finance_org_accepted(
            jab,
            vm_id,
            context,
            expected_text=accepted_text,
            value=value,
            scope_hwnd=scope_hwnd,
        )
        accepted = bool(acceptance_probe.get("accepted"))
        write_ok = bool(paste_result.get("ok") or (set_text_result or {}).get("ok"))
        return {
            "ok": bool(write_ok and accepted),
            "method": "legacy-control-name-guarded-paste-enter",
            "source": "legacy-control-name",
            "control_name": HEADER_SCOPE_ANCHOR_TEXT,
            "path": path,
            "window": window_info,
            "scope_hwnd": scope_hwnd,
            "text_before": before,
            "description_before": (
                info_before.description.strip() if info_before else None
            ),
            "set_ok": write_ok,
            "set_text_ok": bool((set_text_result or {}).get("ok")),
            "guarded_paste": paste_result,
            "set_text_fallback": set_text_result,
            "enter_ok": bool(paste_result.get("enter_ok")),
            "post_write_snapshot": backend_state,
            "acceptance_probe": acceptance_probe,
            "accepted_text": accepted_text_from_backend(
                acceptance_probe.get("matched_snapshot") or backend_state,
                value,
                accepted_text,
            ),
            "text_after": after,
            "description_after": (
                info_after.description.strip() if info_after else None
            ),
        }
    finally:
        jab.release_contexts(vm_id, owned_contexts)


def confirm_finance_org_accepted(
    jab,
    vm_id,
    context,
    expected_text=FINANCE_ORG_ACCEPTED_TEXT,
    value=None,
    scope_hwnd=None,
    timeout=2.0,
    interval=0.15,
):
    deadline = time.time() + max(float(timeout or 0), 0.0)
    attempts = []
    first = True
    scope_probe = None
    while first or time.time() < deadline:
        first = False
        info = jab.get_context_info(vm_id, context)
        text = jab.get_text_context_value(vm_id, context)
        snapshot = describe_backend_field_state(
            info,
            text,
            value=value,
            accepted_text=expected_text,
        )
        attempts.append(snapshot)
        if snapshot.get("accepted"):
            return {
                "ok": True,
                "accepted": True,
                "expected_text": expected_text,
                "source": "current-context",
                "attempts": attempts,
                "matched_snapshot": snapshot,
            }
        if time.time() >= deadline:
            break
        time.sleep(max(float(interval or 0), 0.0))
    scope_probe = probe_finance_org_accepted_text_in_scope(
        jab,
        expected_text,
        scope_hwnd=scope_hwnd,
    )
    if scope_probe.get("accepted"):
        return {
            "ok": True,
            "accepted": True,
            "expected_text": expected_text,
            "source": "scope-text-probe",
            "attempts": attempts,
            "matched_snapshot": scope_probe.get("snapshot"),
            "scope_probe": scope_probe,
        }
    return {
        "ok": False,
        "accepted": False,
        "expected_text": expected_text,
        "attempts": attempts,
        "scope_probe": scope_probe,
        "reason": "财务组织未确认解析为中文",
    }


def probe_finance_org_accepted_text_in_scope(
    jab,
    expected_text=FINANCE_ORG_ACCEPTED_TEXT,
    scope_hwnd=None,
):
    context, vm_id, owned_contexts, owned_indexes, window_info = (
        find_context_with_window(
            jab,
            expected_text,
            roles=(),
            timeout=0.05,
            require_showing=False,
            window_class="SunAwtCanvas",
            visible_only=True,
            scope_hwnd=scope_hwnd,
        )
    )
    if not context:
        return {
            "ok": False,
            "accepted": False,
            "expected_text": expected_text,
            "reason": "scope accepted text not found",
        }
    try:
        info = jab.get_context_info(vm_id, context)
        text = jab.get_text_context_value(vm_id, context)
        snapshot = describe_backend_field_state(
            info,
            text,
            value=None,
            accepted_text=expected_text,
        )
        return {
            "ok": bool(snapshot.get("accepted")),
            "accepted": bool(snapshot.get("accepted")),
            "expected_text": expected_text,
            "path": "0" + "".join(f".{index}" for index in owned_indexes),
            "window": window_info,
            "snapshot": snapshot,
        }
    finally:
        jab.release_contexts(vm_id, owned_contexts)


def guarded_paste_header_value(jab, vm_id, context, window_info, value):
    focus_ok = True
    if hasattr(jab.dll, "requestFocus"):
        focus_ok = bool(jab.dll.requestFocus(vm_id, context))
    if not focus_ok:
        return {
            "ok": False,
            "method": "guarded-clipboard-paste",
            "reason": "JAB requestFocus 失败，未发送剪贴板输入",
            "focus_ok": False,
        }
    guard = foreground_matches_window({"hwnd": (window_info or {}).get("hwnd")})
    if not guard.get("ok"):
        return {
            **guard,
            "ok": False,
            "method": "guarded-clipboard-paste",
            "reason": f"{guard.get('reason')}，未发送表头剪贴板输入",
            "focus_ok": True,
        }
    old_clipboard = get_clipboard_text()
    result = None
    try:
        set_clipboard_text(str(value))
        time.sleep(0.02)
        send_hotkey_ctrl_a()
        time.sleep(0.02)
        send_hotkey_ctrl_v()
        time.sleep(0.02)
        jab.press_key("enter", wait=0)
        result = {
            **guard,
            "ok": True,
            "method": "guarded-clipboard-paste",
            "focus_ok": True,
            "enter_ok": True,
            "enter_method": "jab.press_key",
        }
        return result
    finally:
        try:
            clipboard_restored = restore_clipboard_text(old_clipboard)
        except Exception:
            clipboard_restored = False
        if result is not None:
            result["clipboard_restored"] = clipboard_restored


def find_receipt_header_field_by_semantic_label(
    jab,
    label,
    scope_hwnd=None,
    timeout=1.5,
):
    label_found = find_header_label_context_with_window(
        jab,
        label,
        timeout=timeout,
        require_showing=label != HEADER_SCOPE_ANCHOR_LABEL,
        scope_hwnd=scope_hwnd,
    )
    label_context, vm_id, owned_contexts, owned_indexes, window = label_found
    if not label_context:
        return {"ok": False, "label": label, "reason": "semantic label not found"}
    label_path = None
    if owned_indexes:
        label_path = "0" + "".join(f".{index}" for index in owned_indexes)
    jab.release_contexts(vm_id, owned_contexts)
    if not label_path:
        return {"ok": False, "label": label, "reason": "semantic label path missing"}
    text_path = infer_header_text_path_from_label_path(label, label_path)
    if not text_path:
        return {
            "ok": False,
            "label": label,
            "label_path": label_path,
            "reason": "semantic label path cannot infer text path",
        }
    label_window_hwnd = (window or {}).get("hwnd") or scope_hwnd
    context, vm_id, owned_contexts, window_info = jab.find_context_by_path_once(
        text_path,
        class_name="SunAwtCanvas",
        scope_hwnd=label_window_hwnd,
        role="text",
        require_showing=True,
        require_valid_bounds=False,
    )
    if not context:
        return {
            "ok": False,
            "label": label,
            "label_path": label_path,
            "path": text_path,
            "reason": "semantic inferred text path not found",
            "label_window": window,
            "path_scope_hwnd": label_window_hwnd,
        }
    return {
        "ok": True,
        "label": label,
        "context": context,
        "vm_id": vm_id,
        "owned_contexts": owned_contexts,
        "path": text_path,
        "label_path": label_path,
        "window": window_info,
    }


def find_header_label_context_with_window(
    jab,
    label,
    timeout=1.5,
    require_showing=True,
    scope_hwnd=None,
    strict_anchor=False,
):
    deadline = time.time() + max(float(timeout or 0), 0.0)
    last_window_count = 0
    while time.time() < deadline:
        windows = jab.get_scoped_windows(scope_hwnd, include_children=True)
        last_window_count = len(windows)
        for hwnd, title, class_name, pid, visible in windows:
            if not visible or class_name != "SunAwtCanvas":
                continue
            if not jab.dll.isJavaWindow(hwnd):
                continue
            from tools.jab_probe import JOBJECT

            vm_id_ref = ctypes.c_long()
            root_context = JOBJECT()
            if not jab.dll.getAccessibleContextFromHWND(
                hwnd,
                ctypes.byref(vm_id_ref),
                ctypes.byref(root_context),
            ):
                continue
            context, owned_contexts, owned_indexes = find_header_label_in_tree(
                jab,
                vm_id_ref.value,
                root_context.value,
                label,
                require_showing,
                strict_anchor,
                depth=0,
                owned_contexts=[],
                owned_indexes=[],
            )
            if context:
                return (
                    context,
                    vm_id_ref.value,
                    owned_contexts,
                    owned_indexes,
                    {
                        "hwnd": int(hwnd),
                        "title": title,
                        "class_name": class_name,
                        "pid": pid,
                        "visible": visible,
                    },
                )
            jab.release_contexts(vm_id_ref.value, [root_context.value])
        time.sleep(0.1)
    return None, None, [], [], {"window_count": last_window_count}


def find_header_label_in_tree(
    jab,
    vm_id,
    context,
    label,
    require_showing,
    strict_anchor,
    depth,
    owned_contexts,
    owned_indexes,
):
    info = jab.get_context_info(vm_id, context)
    if not info:
        return None, [], []
    role = (info.role_en_US.strip() or info.role.strip()).lower()
    states = (info.states_en_US.strip() or info.states.strip()).lower()
    if (
        role == "label"
        and (
            header_scope_anchor_text_matches(info)
            if strict_anchor
            else header_label_text_matches(info, label)
        )
        and (not require_showing or ("visible" in states and "showing" in states))
    ):
        return context, list(owned_contexts), list(owned_indexes)
    if depth >= min(jab.max_depth, 50):
        return None, [], []
    if role == "table":
        return None, [], []
    for index in range(min(info.childrenCount, jab.max_children)):
        child = jab.dll.getAccessibleChildFromContext(vm_id, context, index)
        if not child:
            continue
        found, found_contexts, found_indexes = find_header_label_in_tree(
            jab,
            vm_id,
            child,
            label,
            require_showing,
            strict_anchor,
            depth + 1,
            owned_contexts + [child],
            owned_indexes + [index],
        )
        if found:
            return found, found_contexts, found_indexes
        jab.release_contexts(vm_id, [child])
    return None, [], []


def header_label_text_matches(info, label):
    expected = str(label or "").strip()
    if not expected:
        return False
    texts = (
        info.name.strip(),
        info.description.strip(),
    )
    for text in texts:
        if not text:
            continue
        normalized = text.replace("（", "(").replace("）", ")")
        if normalized == expected or normalized.startswith(f"{expected}("):
            return True
    return False


def header_scope_anchor_text_matches(info):
    if not info:
        return False
    for text in (info.name.strip(), info.description.strip()):
        if text.strip() == HEADER_SCOPE_ANCHOR_TEXT:
            return True
    return False


def infer_header_text_path_from_label_path(label, label_path):
    parts = split_header_path(label_path)
    if not parts:
        return None
    if label == "财务组织":
        if parts[-1] == 0:
            return ".".join(str(part) for part in [*parts[:-1], 2, 1, 0])
        return None
    if parts[-1] % 2 != 0:
        return None
    return ".".join(str(part) for part in [*parts[:-1], parts[-1] + 1, 0])


def split_header_path(path):
    try:
        return [int(part) for part in str(path).split(".") if part != ""]
    except ValueError:
        return []


def accepted_text_from_backend(backend_state, raw_value=None, preferred=None):
    if preferred:
        return str(preferred).strip()
    for key in ("description", "text", "name"):
        text = str((backend_state or {}).get(key) or "").strip()
        if text and text != str(raw_value or "").strip():
            return text
    return ""


def build_receipt_header_dynamic_path(dynamic_index, label):
    cached_template = get_receipt_header_path_template(dynamic_index)
    if cached_template:
        path = build_receipt_header_path_from_template(
            dynamic_index,
            label,
            cached_template,
        )
        if path:
            return path
    if label == "财务组织":
        label_path = build_receipt_header_dynamic_label_path(dynamic_index, label)
        return infer_header_text_path_from_label_path(label, label_path)
    index = HEADER_FORM_TEXT_INDEXES.get(label)
    if index is None:
        return None
    suffix = HEADER_COMMON_SUFFIX_TEMPLATE.format(index=index)
    return f"{HEADER_DYNAMIC_PREFIX_BASE}.{dynamic_index}.{suffix}"


def build_receipt_header_dynamic_label_path(dynamic_index, label):
    cached_template = get_receipt_header_path_template(dynamic_index)
    if cached_template:
        path = build_receipt_header_label_path_from_template(
            dynamic_index,
            label,
            cached_template,
        )
        if path:
            return path
    if label == "财务组织":
        return (
            f"{HEADER_DYNAMIC_PREFIX_BASE}.{dynamic_index}.{FINANCE_ORG_LABEL_SUFFIX}"
        )
    index = HEADER_FORM_TEXT_INDEXES.get(label)
    if index is None:
        return None
    suffix = HEADER_COMMON_LABEL_SUFFIX_TEMPLATE.format(index=index - 1)
    return f"{HEADER_DYNAMIC_PREFIX_BASE}.{dynamic_index}.{suffix}"


def receipt_header_dynamic_prefix(dynamic_index):
    return f"{HEADER_DYNAMIC_PREFIX_BASE}.{dynamic_index}"


def receipt_header_default_path_template():
    return {
        "source": "default-header-template",
        "text_suffix_template": HEADER_COMMON_SUFFIX_TEMPLATE,
        "label_suffix_template": HEADER_COMMON_LABEL_SUFFIX_TEMPLATE,
    }


def set_receipt_header_path_template(dynamic_index, template):
    if dynamic_index is None or not template:
        return None
    cache = getattr(set_receipt_header_path_template, "_cache", None)
    if cache is None:
        cache = {}
        setattr(set_receipt_header_path_template, "_cache", cache)
    cache[int(dynamic_index)] = dict(template)
    return cache[int(dynamic_index)]


def get_receipt_header_path_template(dynamic_index):
    if dynamic_index is None:
        return None
    cache = getattr(set_receipt_header_path_template, "_cache", None) or {}
    return cache.get(int(dynamic_index))


def clear_receipt_header_path_template_cache():
    setattr(set_receipt_header_path_template, "_cache", {})


def build_receipt_header_path_from_template(dynamic_index, label, template):
    if label == "财务组织":
        return None
    index = HEADER_FORM_TEXT_INDEXES.get(label)
    suffix_template = (template or {}).get("text_suffix_template")
    if dynamic_index is None or index is None or not suffix_template:
        return None
    return f"{HEADER_DYNAMIC_PREFIX_BASE}.{dynamic_index}.{suffix_template.format(index=index)}"


def build_receipt_header_label_path_from_template(dynamic_index, label, template):
    if label == "财务组织":
        return None
    index = HEADER_FORM_TEXT_INDEXES.get(label)
    suffix_template = (template or {}).get("label_suffix_template")
    if dynamic_index is None or index is None or not suffix_template:
        return None
    return f"{HEADER_DYNAMIC_PREFIX_BASE}.{dynamic_index}.{suffix_template.format(index=index - 1)}"


def infer_header_path_template_from_field(path, dynamic_index, label):
    index = HEADER_FORM_TEXT_INDEXES.get(label)
    if not path or dynamic_index is None or index is None:
        return None
    prefix = f"{HEADER_DYNAMIC_PREFIX_BASE}.{dynamic_index}."
    suffix = str(path)
    if not suffix.startswith(prefix):
        return None
    suffix = suffix[len(prefix) :]
    marker = f".{index}.0"
    if not suffix.endswith(marker):
        return None
    base = suffix[: -len(marker)]
    if not base:
        return None
    return {
        "source": f"learned-from-{label}",
        "text_suffix_template": f"{base}.{{index}}.0",
        "label_suffix_template": f"{base}.{{index}}",
        "sample_label": label,
        "sample_path": path,
    }


def extract_receipt_header_dynamic_index(path):
    prefix = f"{HEADER_DYNAMIC_PREFIX_BASE}."
    if not path or not path.startswith(prefix):
        return None
    first = path[len(prefix) :].split(".", 1)[0]
    try:
        return int(first)
    except ValueError:
        return None


def find_receipt_header_field_by_dynamic_path(
    jab,
    label,
    dynamic_index,
    scope_hwnd=None,
    require_showing=True,
    require_valid_bounds=True,
    path_template=None,
):
    text_path = (
        build_receipt_header_path_from_template(dynamic_index, label, path_template)
        if path_template
        else build_receipt_header_dynamic_path(dynamic_index, label)
    )
    if not text_path:
        return {
            "ok": False,
            "reason": "header dynamic path not configured",
            "label": label,
            "dynamic_index": dynamic_index,
        }
    context, vm_id, owned_contexts, window_info = jab.find_context_by_path_once(
        text_path,
        class_name="SunAwtCanvas",
        scope_hwnd=scope_hwnd,
        role="text",
        require_showing=require_showing,
        require_valid_bounds=require_valid_bounds,
    )
    if not context:
        return {
            "ok": False,
            "reason": "header dynamic path not found",
            "label": label,
            "path": text_path,
            "dynamic_index": dynamic_index,
        }
    return {
        "ok": True,
        "context": context,
        "vm_id": vm_id,
        "owned_contexts": owned_contexts,
        "path": text_path,
        "label_path": (
            build_receipt_header_label_path_from_template(
                dynamic_index,
                label,
                path_template,
            )
            if path_template
            else build_receipt_header_dynamic_label_path(dynamic_index, label)
        ),
        "window": window_info,
        "dynamic_index": dynamic_index,
        "dynamic_prefix": receipt_header_dynamic_prefix(dynamic_index),
        "path_template": path_template,
    }


def find_receipt_header_field_by_scoped_label(jab, label, scope_hwnd=None):
    aliases = HEADER_LABEL_ALIASES.get(label, (label,))
    for hwnd, title, class_name, pid, visible in jab.get_scoped_windows(
        scope_hwnd, include_children=True
    ):
        if not visible or class_name != "SunAwtCanvas":
            continue
        if scope_hwnd is not None and int(hwnd) != int(scope_hwnd):
            continue
        if not jab.dll.isJavaWindow(hwnd):
            continue
        from tools.jab_probe import JOBJECT

        vm_id_ref = ctypes.c_long()
        root_context = JOBJECT()
        if not jab.dll.getAccessibleContextFromHWND(
            hwnd,
            ctypes.byref(vm_id_ref),
            ctypes.byref(root_context),
        ):
            continue
        for alias in aliases:
            result = find_label_following_text(
                jab,
                vm_id_ref.value,
                root_context.value,
                alias,
                path="0",
                depth=0,
                owned_contexts=[root_context.value],
            )
            if result:
                context, owned_contexts, path, label_path = result
                return {
                    "ok": True,
                    "label": label,
                    "matched_alias": alias,
                    "context": context,
                    "vm_id": vm_id_ref.value,
                    "owned_contexts": owned_contexts,
                    "path": path,
                    "label_path": label_path,
                    "window": {
                        "hwnd": int(hwnd),
                        "title": title,
                        "class": class_name,
                        "pid": pid,
                        "visible": visible,
                    },
                    "dynamic_index": extract_receipt_header_dynamic_index(path),
                    "dynamic_prefix": receipt_header_dynamic_prefix(
                        extract_receipt_header_dynamic_index(path)
                    )
                    if extract_receipt_header_dynamic_index(path) is not None
                    else None,
                    "source": "scoped-label-following-text",
                }
        jab.release_contexts(vm_id_ref.value, [root_context.value])
    return {
        "ok": False,
        "label": label,
        "reason": "scoped label-following text not found",
        "scope_hwnd": scope_hwnd,
        "aliases": aliases,
    }


def find_label_following_text(jab, vm_id, context, label, path, depth, owned_contexts):
    info = jab.get_context_info(vm_id, context)
    if not info:
        return None
    role = (info.role_en_US.strip() or info.role.strip()).lower()
    if depth >= jab.max_depth or role == "table":
        return None

    child_infos = []
    for index in range(min(info.childrenCount, jab.max_children)):
        child = jab.dll.getAccessibleChildFromContext(vm_id, context, index)
        if not child:
            continue
        child_info = jab.get_context_info(vm_id, child)
        child_path = f"{path}.{index}"
        if child_info:
            child_infos.append((index, child, child_info, child_path))
        else:
            jab.release_contexts(vm_id, [child])

    for position, (_index, child, child_info, child_path) in enumerate(child_infos):
        child_role = (child_info.role_en_US.strip() or child_info.role.strip()).lower()
        child_states = (
            child_info.states_en_US.strip() or child_info.states.strip()
        ).lower()
        child_texts = {child_info.name.strip(), child_info.description.strip()}
        if child_role == "label" and label in child_texts and "visible" in child_states:
            for _next_index, next_child, _next_info, next_path in child_infos[
                position + 1 :
            ]:
                text_context, text_owned, text_path = first_text_descendant(
                    jab,
                    vm_id,
                    next_child,
                    next_path,
                    depth + 1,
                )
                if text_context:
                    keep = [item[1] for item in child_infos] + text_owned
                    return text_context, owned_contexts + keep, text_path, child_path

    for _index, child, _child_info, child_path in child_infos:
        result = find_label_following_text(
            jab,
            vm_id,
            child,
            label,
            child_path,
            depth + 1,
            owned_contexts + [item[1] for item in child_infos],
        )
        if result:
            return result
    for _index, child, _child_info, _child_path in child_infos:
        jab.release_contexts(vm_id, [child])
    return None


def first_text_descendant(jab, vm_id, context, path, depth):
    info = jab.get_context_info(vm_id, context)
    if not info:
        return None, [], None
    role = (info.role_en_US.strip() or info.role.strip()).lower()
    states = (info.states_en_US.strip() or info.states.strip()).lower()
    if role == "text" and "visible" in states and "editable" in states:
        return context, [], path
    if depth >= jab.max_depth or role == "table":
        return None, [], None
    owned = []
    for index in range(min(info.childrenCount, jab.max_children)):
        child = jab.dll.getAccessibleChildFromContext(vm_id, context, index)
        if not child:
            continue
        child_path = f"{path}.{index}"
        found, found_owned, found_path = first_text_descendant(
            jab,
            vm_id,
            child,
            child_path,
            depth + 1,
        )
        if found:
            return found, owned + [child] + found_owned, found_path
        jab.release_contexts(vm_id, [child])
    return None, owned, None


def locate_receipt_header_scope(jab, scope_hwnd=None):
    semantic = infer_receipt_header_scope_by_semantic(jab, scope_hwnd=scope_hwnd)
    if semantic.get("ok"):
        return semantic
    return {
        "ok": False,
        "semantic_attempt": semantic,
        "reason": "未能用语义路径推断当前收款单表头 scope",
    }


def infer_receipt_header_scope_by_semantic(jab, scope_hwnd=None):
    found = find_receipt_header_field_by_semantic_label(
        jab,
        HEADER_SCOPE_ANCHOR_LABEL,
        scope_hwnd=scope_hwnd,
    )
    if not found.get("ok"):
        return {
            "ok": False,
            "mode": "semantic-path-inference",
            "reason": found.get("reason") or "财务组织语义定位失败",
            "attempt": found,
        }
    dynamic_index = extract_receipt_header_dynamic_index(found.get("path"))
    scope_hwnd = ((found.get("window") or {}).get("hwnd")) or None
    jab.release_contexts(found["vm_id"], found["owned_contexts"])
    if dynamic_index is None or not scope_hwnd:
        return {
            "ok": False,
            "mode": "semantic-path-inference",
            "reason": "财务组织语义路径无法推出动态前缀或窗口",
            "attempt": {
                "path": found.get("path"),
                "label_path": found.get("label_path"),
                "window": found.get("window"),
            },
        }
    return {
        "ok": True,
        "scope_hwnd": scope_hwnd,
        "mode": "semantic-path-inference",
        "dynamic_index": dynamic_index,
        "dynamic_prefix": receipt_header_dynamic_prefix(dynamic_index),
        "matched_labels": [HEADER_SCOPE_ANCHOR_LABEL],
        "semantic_label_path": found.get("label_path"),
        "semantic_text_path": found.get("path"),
    }


def foreground_root_hwnd():
    if os.name != "nt" or not hasattr(ctypes, "windll"):
        return 0
    hwnd = ctypes.windll.user32.GetForegroundWindow()
    return window_root_hwnd(hwnd)


def window_root_hwnd(hwnd):
    if os.name != "nt" or not hasattr(ctypes, "windll") or not hwnd:
        return 0
    return int(ctypes.windll.user32.GetAncestor(wintypes.HWND(int(hwnd)), 2) or 0)


def find_context_with_window(
    jab,
    name,
    roles=(),
    timeout=None,
    require_showing=False,
    window_title=None,
    window_class=None,
    visible_only=True,
    scope_hwnd=None,
):
    deadline = time.time() + (timeout or jab.search_timeout)
    normalized_roles = {role.lower() for role in roles}
    while time.time() < deadline:
        windows = jab.get_scoped_windows(scope_hwnd, include_children=True)
        for hwnd, title, class_name, pid, visible in windows:
            if visible_only and not visible:
                continue
            if (
                scope_hwnd is None
                and window_title is not None
                and title != window_title
            ):
                continue
            if (
                scope_hwnd is None
                and window_class is not None
                and class_name != window_class
            ):
                continue
            if not jab.dll.isJavaWindow(hwnd):
                continue
            from tools.jab_probe import JOBJECT

            vm_id_ref = ctypes.c_long()
            root_context = JOBJECT()
            if not jab.dll.getAccessibleContextFromHWND(
                hwnd,
                ctypes.byref(vm_id_ref),
                ctypes.byref(root_context),
            ):
                continue
            context, owned_contexts, owned_indexes = jab.find_in_tree_with_path(
                vm_id_ref.value,
                root_context.value,
                name,
                normalized_roles,
                require_showing,
                depth=0,
                owned_contexts=[],
                owned_indexes=[],
            )
            if context:
                return (
                    context,
                    vm_id_ref.value,
                    owned_contexts,
                    owned_indexes,
                    {
                        "hwnd": int(hwnd),
                        "title": title,
                        "class_name": class_name,
                        "pid": pid,
                        "visible": visible,
                    },
                )
            jab.release_contexts(vm_id_ref.value, [root_context.value])
        time.sleep(0.2)
    return None, None, [], [], {}


def post_key_to_hwnd(hwnd, key):
    if os.name != "nt" or not hwnd:
        return False
    key_map = {
        "enter": 0x0D,
        "tab": 0x09,
    }
    vk = key_map.get(str(key).lower())
    if not vk:
        return False
    user32 = ctypes.windll.user32
    hwnd = wintypes.HWND(int(hwnd))
    WM_KEYDOWN = 0x0100
    WM_KEYUP = 0x0101
    down_ok = bool(user32.PostMessageW(hwnd, WM_KEYDOWN, vk, 0))
    up_ok = bool(user32.PostMessageW(hwnd, WM_KEYUP, vk, 0))
    return down_ok and up_ok


def do_context_commit_action(jab, vm_id, context):
    actions = jab.get_action_names(vm_id, context)
    preferred = ("确认", "确定", "提交", "单击", "click", "press")
    for action_name in preferred:
        if action_name not in actions:
            continue
        try:
            ok = bool(jab.do_action(vm_id, context, action_name=action_name))
        except Exception as exc:
            return {
                "ok": False,
                "action": action_name,
                "exception": repr(exc),
                "actions": actions,
            }
        if ok:
            return {"ok": True, "action": action_name, "actions": actions}
    return {"ok": False, "reason": "no commit action", "actions": actions}


def describe_backend_field_state(info, text, value=None, accepted_text=None):
    name = info.name.strip() if info else ""
    description = info.description.strip() if info else ""
    accepted = backend_field_accepts(info, text, value, accepted_text)
    written = backend_field_has_written_value(info, text, value)
    return {
        "accepted": bool(accepted),
        "written": bool(written),
        "unlocked": False,
        "text": text,
        "name": name,
        "description": description,
    }


def backend_field_has_written_value(info, text, value=None):
    expected = str(value).strip() if value is not None else ""
    if not info or not expected:
        return False
    actual_text = str(text or "").strip()
    description = info.description.strip()
    return actual_text == expected or description == expected


def backend_field_accepts(info, text, value=None, accepted_text=None):
    if not info:
        return False
    if accepted_text:
        return context_contains(info, accepted_text)
    expected = str(value).strip() if value is not None else ""
    actual_text = str(text or "").strip()
    description = info.description.strip()
    if expected and (actual_text == expected or description == expected):
        return True
    return bool(description)


def context_contains(info, expected_text):
    if not info or not expected_text:
        return False
    expected = str(expected_text).strip()
    haystack = " ".join(
        part
        for part in (
            info.name.strip(),
            info.description.strip(),
            info.role.strip(),
            info.role_en_US.strip(),
            info.states.strip(),
            info.states_en_US.strip(),
        )
        if part
    )
    return expected in haystack


def fill_detail_line(jab, business):
    steps = []
    for col, name, value in [
        (1, "收款业务类型", business["main_business_type"]),
        (4, "收款银行账户", business["bank_account"]),
        (5, "科目", business["main_subject"]),
        (7, "贷方原币金额", business["amount"]),
        (11, "结算方式", business["settlement"]),
    ]:
        steps.append(fill_cell(jab, 0, col, name, value))
    return steps


def fill_cell(jab, row, col, name, value):
    attempts = []
    for mode, kwargs in [
        (
            "direct",
            {"set_cell_text": value, "commit_key": "none", "focus_target": "cell"},
        ),
    ]:
        try:
            result = select_cell(
                jab,
                table_index=0,
                row=row,
                col=col,
                window_title=None,
                locate_body_table=True,
                wait=0.35,
                **kwargs,
            )
        except Exception as exc:
            result = {"ok": False, "exception": repr(exc)}
        attempts.append({"mode": mode, "result": compact_selection_result(result)})
        if cell_changed(result, value):
            break
        time.sleep(0.2)
    ok = any(cell_changed(attempt["result"], value) for attempt in attempts)
    return {
        "step": "detail_cell",
        "ok": ok,
        "blocked": not ok,
        "reason": (
            None
            if ok
            else "backend cell input failed; global keyboard input is disabled"
        ),
        "row": row,
        "col": col,
        "name": name,
        "value": value,
        "attempts": attempts,
    }


def compact_selection_result(result):
    if not isinstance(result, dict):
        return result
    keep = {
        key: result.get(key)
        for key in (
            "ok",
            "reason",
            "row_count",
            "col_count",
            "child_index",
            "selected_before",
            "selected_after",
            "cell_text_before",
            "cell_text_after",
            "edit",
            "exception",
        )
        if key in result
    }
    return keep


def cell_changed(result, value):
    after = str((result or {}).get("cell_text_after") or "")
    return (
        after
        and after != str((result or {}).get("cell_text_before") or "")
        and str(value) in after
    )


def read_body_table(jab, step, scope_hwnd=None):
    located = locate_receipt_body_table(jab, max_rows=3, scope_hwnd=scope_hwnd)
    best = located.get("best")
    if not best:
        return {
            "step": step,
            "ok": False,
            "reason": "body table not found",
            "candidates": located.get("candidates", [])[:3],
        }
    return {
        "step": step,
        "ok": True,
        "path": best.get("path"),
        "row_count": best.get("row_count"),
        "col_count": best.get("col_count"),
        "rows": best.get("rows"),
    }


if __name__ == "__main__":
    raise SystemExit(main())
