# 生命周期：T0 一次性（删除条件：授权主体真实保存+Sheet2写回验证完成后删除）
# 覆盖的业务阶段：收款单自制录入-授权主体真实保存-Sheet2结果追加-后验查询
# 依赖的服务/环境：Windows Python、NC 收款单录入页、Java Access Bridge、收款单Excel
# 运行方式：python tools/tmp_receipt_three_org_real_save_run.py

from collections import Counter
from copy import deepcopy
from datetime import date
import os
import sys
import time
import traceback
from dataclasses import replace
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from core.receipt_entry import (  # noqa: E402
    RESULT_SHEET_HEADERS,
    ReceiptEntryWorkbook,
    append_plan_sheet_row,
    ensure_result_sheet_headers,
    names_match,
    parse_amount,
    plan_sheet_row,
)
from core.utils import load_config  # noqa: E402
from tools import tmp_receipt_two_case_save_run as base_run  # noqa: E402
from tools.receipt_query_fill import fill_receipt_query  # noqa: E402


AUTHORIZED_ORGS = ("A001", "A006")
QUERY_MAX_ROWS = 500
ALREADY_SAVED_ROWS = {
    int(item)
    for item in os.environ.get("RECEIPT_ALREADY_SAVED_ROWS", "").split(",")
    if item.strip()
}


def elapsed(start):
    return round(time.perf_counter() - start, 3)


def choose_cases(config):
    workbook = ReceiptEntryWorkbook(config)
    rows, issues, _summary = workbook.build_local_plan(write_sheet=False)
    issue_rows = {issue.excel_row for issue in issues if issue.excel_row is not None}
    valid_rows = [row for row in rows if row.row not in issue_rows]

    selected: list[tuple[object, str, str]] = []
    used_rows = set()

    def pick_candidate(predicate, missing_message, tag, fee_mode):
        candidates = [
            row
            for row in valid_rows
            if predicate(row)
            and row.row not in used_rows
            and row.row not in ALREADY_SAVED_ROWS
        ]
        candidates.sort(key=lambda row: (row.receipt_date, row.row), reverse=True)
        if not candidates:
            raise RuntimeError(missing_message)
        row = candidates[0]
        selected.append((row, tag, fee_mode))
        used_rows.add(row.row)

    cny_candidates = [
        row
        for row in valid_rows
        if row.organization_code == "A001"
        and row.header_currency_code == "CNY"
        and row.row not in used_rows
        and row.row not in ALREADY_SAVED_ROWS
    ]
    cny_candidates.sort(key=lambda row: (row.receipt_date, row.row), reverse=True)
    if not cny_candidates:
        raise RuntimeError("找不到 A001/CNY 有效测试行")
    selected.append((cny_candidates[0], "cny", "source"))
    used_rows.add(cny_candidates[0].row)

    pick_candidate(
        lambda row: row.organization_code == "A006"
        and row.header_currency_code == "USD",
        "找不到 A006/USD 有效测试行",
        "usd-fee",
        "force_fee",
    )
    pick_candidate(
        lambda row: row.organization_code in AUTHORIZED_ORGS
        and row.header_currency_code == "USD",
        "找不到授权主体 USD 无手续费有效测试行",
        "usd-no-fee",
        "force_no_fee",
    )

    # 覆盖：授权主体、CNY、USD、有手续费、无手续费。CNY 保持源手续费。
    cases = []
    for index, (row, tag, fee_mode) in enumerate(selected, start=1):
        fee = str(row.fee)
        if fee_mode == "force_fee":
            fee = "20.00"
        elif fee_mode == "force_no_fee":
            fee = "0.00"
        cases.append(
            base_run.TestCase(
                name=f"授权主体真实保存{index}-{row.organization_short_name}-{tag}",
                excel_row=row.row,
                document_date=row.receipt_date.isoformat(),
                customer_code=row.customer_code,
                payer_name=row.payer_name,
                source_bank=row.bank,
                bank_label=row.account_label,
                bank_account_no=row.account_no,
                currency=row.currency,
                amount=str(row.raw_amount),
                fee=fee,
            )
        )
    return cases


def append_case_result_to_sheet2(config, case_report, query_result):
    workbook = ReceiptEntryWorkbook(config)
    rows, issues, _summary = workbook.build_local_plan(write_sheet=False)
    row_by_number = {row.row: row for row in rows}
    row = row_by_number.get(case_report.get("excel_row"))
    if row is None:
        raise RuntimeError(f"找不到原 Sheet1 行: {case_report.get('excel_row')}")
    business = case_report.get("business") or {}
    if business.get("fee") is not None:
        row = replace(row, fee=parse_amount(business["fee"]))

    local_status = "通过" if case_report.get("ok") else "异常"
    values = plan_sheet_row(row, None, local_status)
    row_by_header = dict(zip(RESULT_SHEET_HEADERS, values, strict=True))
    row_by_header["录入结果"] = format_entry_result(case_report)
    row_by_header["保存结果"] = format_save_result(case_report)
    row_by_header["后验查询结果"] = query_result

    wb = openpyxl.load_workbook(workbook.excel_path, read_only=False)
    try:
        name = workbook.config.result_sheet_name
        ws = wb[name] if name in wb.sheetnames else wb.create_sheet(name)
        columns = ensure_result_sheet_headers(ws, workbook.config.header_row)
        append_plan_sheet_row(
            ws,
            columns,
            max(ws.max_row + 1, workbook.config.header_row + 1),
            [row_by_header[header] for header in RESULT_SHEET_HEADERS],
        )
        wb.save(workbook.excel_path)
    finally:
        wb.close()


def format_entry_result(case_report):
    if case_report.get("ok") or case_report.get("save"):
        return f"录入通过；耗时={case_report.get('seconds')}s"
    return (
        f"录入失败；阶段={case_report.get('failed_step')}; "
        f"原因={case_report.get('reason')}; 耗时={case_report.get('seconds')}s"
    )


def format_save_result(case_report):
    save = case_report.get("save") or {}
    if not save:
        return f"未保存；原因={case_report.get('reason')}"
    if save.get("ok"):
        return f"保存成功；等待新增成功；耗时={save.get('seconds')}s"
    return f"保存失败；原因={save.get('reason')}; 耗时={save.get('seconds')}s"


def run_post_query(config, case_report):
    if not case_report.get("ok"):
        return "未查询：保存未成功"
    business = case_report.get("business") or {}
    start = time.perf_counter()
    try:
        result = fill_receipt_query(
            deepcopy(config),
            org_code=business["finance_org_code"],
            date_from=business["document_date"],
            date_to=business["document_date"],
            confirm=True,
            read_results=True,
            dry_run_match=False,
            max_rows=QUERY_MAX_ROWS,
        )
        rows = result.get("nc_rows") or []
        amount = parse_amount(business["amount"])
        payer_name = business.get("payer_name")
        matches = [
            row
            for row in rows
            if row.original_amount == amount and names_match(payer_name, row.customer)
        ]
        timing_text = ", ".join(
            f"{item['name']}={item['seconds']}s" for item in result.get("timings") or []
        )
        if len(matches) == 1:
            return (
                f"查询匹配成功；NC行={matches[0].row_index}; "
                f"耗时={elapsed(start)}s; {timing_text}"
            )
        if matches:
            return f"查询重复匹配{len(matches)}条；耗时={elapsed(start)}s; {timing_text}"
        amount_hits = [row for row in rows if row.original_amount == amount]
        return (
            f"查询未唯一匹配；金额命中={len(amount_hits)}; 结果行={len(rows)}; "
            f"耗时={elapsed(start)}s; {timing_text}"
        )
    except Exception as exc:
        return f"查询异常：{type(exc).__name__}: {exc}; 耗时={elapsed(start)}s"


def print_selected(cases):
    print("授权主体真实保存测试数据：")
    for index, case in enumerate(cases, start=1):
        print(
            f"{index}. {case.name} | Sheet1行={case.excel_row} | 日期={case.document_date} "
            f"| 客户={case.customer_code} | 主体银行={case.bank_label} "
            f"| 账号={case.bank_account_no} | 币种={case.currency} "
            f"| 金额={case.amount} | 手续费={case.fee}"
        )
    print("币种覆盖检查：", dict(Counter(case.currency for case in cases)))
    print("主体覆盖检查：", dict(Counter(case.name.split('-')[1] for case in cases)))
    print()


def main():
    run_start = time.perf_counter()
    base_run.SAVE_ENABLED = True
    base_run.TEST_BANK_ACCOUNT_NO = ""
    base_run.ALLOW_EXISTING_ENTRY_FOR_FIRST_CASE = False
    config = load_config(str(ROOT / "config.json"))
    cases = choose_cases(config)
    print_selected(cases)
    print(
        f"请在 {base_run.START_DELAY_SECONDS} 秒内切到 NC【收款单录入】且能看到【新增】的页面..."
    )
    time.sleep(base_run.START_DELAY_SECONDS)

    reports = []
    print("阶段1：连续保存，不做查询；单条失败只记录，继续下一条。")
    for index, case in enumerate(cases):
        if base_run.is_stop_hotkey_pressed():
            print(f"检测到紧急停止键 {base_run.STOP_HOTKEY}，停止。")
            break
        case_start = time.perf_counter()
        report = base_run.run_one_case(config, case, allow_existing_entry=False)
        report["outer_seconds"] = elapsed(case_start)
        base_run.print_case_summary(report)
        reports.append(report)
        if index < len(cases) - 1:
            time.sleep(0.8)

    print()
    print("阶段2：统一后验查询并写 Sheet2。")
    for report in reports:
        if not report.get("ok"):
            print(
                f"  Sheet1行={report.get('excel_row')} | "
                "保存未成功，跳过后验查询，Sheet2 追加失败结果"
            )
            report["post_query_result"] = "未查询：保存未成功"
            append_case_result_to_sheet2(config, report, report["post_query_result"])
            print("  Sheet2：已追加本案例失败结果")
            continue
        query_result = run_post_query(config, report)
        report["post_query_result"] = query_result
        append_case_result_to_sheet2(config, report, query_result)
        print(
            f"  Sheet1行={report.get('excel_row')} | "
            f"后验查询结果：{query_result}"
        )
        print("  Sheet2：已追加本案例结果")

    print()
    print("授权主体真实保存总结果：")
    ok = len(reports) == len(cases) and all(report.get("ok") for report in reports)
    print(f"  {'成功' if ok else '失败'}")
    print(f"  完成案例：{len(reports)}/{len(cases)}")
    print(f"  保存案例累计耗时：{round(sum(float(r.get('seconds') or 0) for r in reports), 3)} 秒")
    print(f"  脚本总耗时：{elapsed(run_start)} 秒")
    for report in reports:
        business = report.get("business") or {}
        print(
            f"  - Sheet1行={report.get('excel_row')} 主体={business.get('finance_org_code')} "
            f"币种={business.get('header_currency_code')} 手续费={business.get('fee')} "
            f"保存={'成功' if report.get('ok') else '失败'} "
            f"耗时={report.get('seconds')}s"
        )
    return 0 if ok else 1


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception:
        traceback.print_exc()
        raise
