import argparse
import json
import subprocess
import sys
import time
from datetime import datetime
from decimal import Decimal
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from core.jab_operator import JABOperator  # noqa: E402
from core.receipt_config import ReceiptEntryConfig  # noqa: E402
from core.utils import load_config  # noqa: E402
from tools.read_receipt_excel_row import DEFAULT_FIELDS  # noqa: E402
# read_body_table 已归位到 receipt_detail_reader(它是读明细表的正式模块);此处 re-export 以保留可导入面与 monkeypatch 锚点
from tools.receipt_detail_reader import read_body_table  # noqa: E402, F401
from tools.receipt_new_probe import (  # noqa: E402
    collect_receipt_new_windows,
    detect_self_made_entry_state,
)
from tools.receipt_keyboard_utils import (  # noqa: E402, F401
    foreground_matches_window,
    get_clipboard_text,
    restore_clipboard_text,
    send_hotkey_ctrl_a,
    send_hotkey_ctrl_v,
    set_clipboard_text,
)
from tools.receipt_table_cell_probe import select_cell  # noqa: E402

# 收款单表头库已纯搬移到 receipt_header_{paths,tree,scope,writer} 与
# receipt_customer_readback；本文件保留探针 CLI(main)与 run_receipt_new_probe[_with_jab]，
# 并 re-export 搬出的名字以保持原可导入面与测试 monkeypatch 语义(tests/外部模块仍从
# tools.receipt_self_made_flow import)。子模块在调用期经各自的 _trial 代理读本模块属性，
# 因此对本模块上这些名字的 monkeypatch 仍对子模块内部协作调用生效。
from tools.receipt_header_paths import (  # noqa: E402, F401
    CURRENCY_NAMES,
    FINANCE_ORG_ACCEPTED_TEXT,
    FINANCE_ORG_LABEL_SUFFIX,
    HEADER_COMMON_LABEL_SUFFIX_TEMPLATE,
    HEADER_COMMON_SUFFIX_TEMPLATE,
    HEADER_DYNAMIC_PREFIX_BASE,
    HEADER_FORM_TEXT_INDEXES,
    HEADER_LIVE_SEMANTIC_FALLBACK_TIMEOUT,
    HEADER_PROBE_LABEL_KEYS,
    HEADER_REQUIRED_LABELS,
    HEADER_SCOPE_ANCHOR_LABEL,
    HEADER_SCOPE_ANCHOR_TEXT,
    accepted_text_from_backend,
    build_receipt_header_dynamic_label_path,
    build_receipt_header_dynamic_path,
    build_receipt_header_label_path_from_template,
    build_receipt_header_path_from_template,
    clear_receipt_header_path_template_cache,
    extract_receipt_header_dynamic_index,
    get_receipt_header_path_template,
    header_label_text_matches,
    header_scope_anchor_text_matches,
    infer_header_path_template_from_field,
    infer_header_text_path_from_label_path,
    receipt_header_default_path_template,
    receipt_header_dynamic_prefix,
    set_receipt_header_path_template,
    split_header_path,
)
from tools.receipt_header_tree import (  # noqa: E402, F401
    find_context_with_window,
    find_header_label_context_with_window,
    find_header_label_in_tree,
    find_label_following_text,
    first_text_descendant,
)
from tools.receipt_header_scope import (  # noqa: E402, F401
    correct_header_anchor_dynamic_index_by_customer,
    infer_receipt_header_scope_by_semantic,
    locate_receipt_header_scope,
    resolve_receipt_header_anchor_in_canvas,
    resolve_receipt_header_scope,
    validate_receipt_header_scope_anchor,
)
from tools.receipt_header_writer import (  # noqa: E402, F401
    backend_field_accepts,
    backend_field_has_written_value,
    confirm_finance_org_accepted,
    context_contains,
    describe_backend_field_state,
    fill_header,
    find_receipt_header_field_by_dynamic_path,
    find_receipt_header_field_by_live_semantic,
    find_receipt_header_field_by_semantic_label,
    guarded_paste_header_value,
    probe_finance_org_accepted_text_in_scope,
    set_finance_org_by_legacy_control_name,
    set_receipt_header_dynamic_field,
    wait_header_account_description,
)
from tools.receipt_customer_readback import (  # noqa: E402, F401
    collect_context_text_rows,
    collect_customer_field_candidates_for_scope,
    context_text_snapshot,
    dedupe_customer_candidates,
    first_valid_customer_name,
    is_valid_customer_name_candidate,
    snapshot_header_field_candidate,
    valid_customer_values_from_snapshot,
)


def print_json(data):
    text = json.dumps(data, ensure_ascii=False, indent=2, default=str)
    try:
        sys.stdout.write(text + "\n")
    except UnicodeEncodeError:
        sys.stdout.buffer.write((text + "\n").encode("utf-8"))


def main():
    parser = argparse.ArgumentParser(
        description="Trial-fill NC receipt self-made entry from one Excel row."
    )
    parser.add_argument("row", type=int, nargs="?")
    parser.add_argument("--config", default="config.json")
    parser.add_argument("--open-self-made", action="store_true")
    parser.add_argument(
        "--probe-header-path-transitions",
        action="store_true",
        help=(
            "只读探测表头 path：当前状态一次，回车后一次，再回车后一次；"
            "用于人工在两次回车之间写财务组织、加窗口"
        ),
    )
    parser.add_argument(
        "--probe-header-semantic-field",
        default=None,
        help="只读测速单个表头字段语义定位，例如 客户、单据日期、币种、结算方式。",
    )
    parser.add_argument(
        "--probe-customer-name-readback",
        action="store_true",
        help="只读探测客户回车后的 NC 客户名称候选来源；不输入、不保存。",
    )
    parser.add_argument(
        "--probe-timeout",
        type=float,
        default=HEADER_LIVE_SEMANTIC_FALLBACK_TIMEOUT,
        help="单次语义定位超时秒数，默认使用正式兜底超时。",
    )
    parser.add_argument(
        "--probe-repeat",
        type=int,
        default=1,
        help="重复测速次数，默认 1。",
    )
    parser.add_argument(
        "--fill-detail",
        action="store_true",
        help="fill receipt detail cells after header verification",
    )
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args()

    config = load_config(args.config)
    if args.probe_header_path_transitions:
        report = probe_header_path_transitions(config)
        print_json(report)
        return 0 if report.get("ok") else 1
    if args.probe_header_semantic_field:
        report = probe_header_semantic_field_speed(
            config,
            normalize_header_probe_label(args.probe_header_semantic_field),
            timeout=args.probe_timeout,
            repeat=args.probe_repeat,
        )
        print_json(report)
        return 0 if report.get("ok") else 1
    if args.probe_customer_name_readback:
        report = probe_customer_name_readback(config, timeout=args.probe_timeout)
        print_json(report)
        return 0 if report.get("ok") else 1
    if args.row is None:
        parser.error(
            "row is required unless --probe-header-path-transitions "
            "or --probe-header-semantic-field or --probe-customer-name-readback is used"
        )
    row_data = read_excel_row(config, args.row)
    business = build_business_values(config, row_data)
    report = {"row": args.row, "excel": row_data, "business": business, "steps": []}

    if args.open_self_made:
        open_step = detect_existing_self_made_entry(config)
        if not open_step.get("ok"):
            open_step = run_receipt_new_probe()
        report["steps"].append(open_step)
        if not open_step.get("ok"):
            report["stopped"] = "open_self_made"
            print_json(report)
            return 1

    jab = JABOperator(config)
    try:
        jab.ensure_started()
        header_steps = fill_header(jab, business)
        report["steps"].extend(header_steps)
        if any(step.get("step") == "blocked" for step in header_steps):
            report["stopped"] = "header"
            print_json(report)
            return 1
        report["steps"].append(read_body_table(jab, "before_detail_fill"))
        if args.fill_detail:
            report["steps"].extend(fill_detail_line(jab, business))
            report["steps"].append(read_body_table(jab, "after_detail_fill"))
        else:
            report["steps"].append(
                {
                    "step": "detail_fill",
                    "ok": False,
                    "blocked": True,
                    "reason": "detail fill requires explicit --fill-detail",
                }
            )
    finally:
        jab.close()

    print_json(report)
    return 0


def read_excel_row(config, row):
    import openpyxl

    excel_cfg = config["receipt_entry"]["excel"]
    workbook = openpyxl.load_workbook(excel_cfg["path"], data_only=True, read_only=True)
    try:
        sheet = workbook[excel_cfg["sheet_name"]]
        header_row = excel_cfg.get("header_row", 1)
        headers = [
            sheet.cell(header_row, col).value for col in range(1, sheet.max_column + 1)
        ]
        data = {
            headers[col - 1]: sheet.cell(row, col).value
            for col in range(1, sheet.max_column + 1)
            if headers[col - 1]
        }
    finally:
        workbook.close()
    return {field: data.get(field) for field in DEFAULT_FIELDS}


def build_business_values(config, row_data):
    receipt_config = ReceiptEntryConfig(config)
    bank = str(row_data.get("银行") or "").strip()
    account = receipt_config.accounts_by_label.get(
        bank.upper()
    ) or receipt_config.accounts_by_label.get(bank)
    if not account:
        normalized_bank = "".join(
            ch for ch in bank.upper() if ch.isalnum() or "\u4e00" <= ch <= "\u9fff"
        )
        account = receipt_config.accounts_by_label.get(normalized_bank)
    if not account:
        raise SystemExit(f"bank account config not found: {bank!r}")
    organization = receipt_config.organizations[account.organization_code]
    currency_code = str(row_data.get("币种") or "").strip().upper()
    receipt_date = row_data.get("到款日期")
    if isinstance(receipt_date, datetime):
        receipt_date_text = receipt_date.strftime("%Y-%m-%d")
    else:
        receipt_date_text = str(receipt_date)[:10]
    amount = Decimal(str(row_data.get("🟪到账金额") or "0"))
    fee_raw = row_data.get("手续费")
    fee = Decimal(str(fee_raw or "0"))
    return {
        "finance_org_code": organization.code,
        "finance_org_name": organization.name,
        "document_date": receipt_date_text,
        "customer_code": str(row_data.get("客户编码") or "").strip(),
        "currency": CURRENCY_NAMES.get(currency_code, currency_code),
        "header_currency_code": account.header_currency_code,
        "bank_label": bank,
        "bank_account": account.account_no,
        "amount": str(amount),
        "fee": str(fee),
        "has_fee": fee != 0,
        "settlement": "网银",
        "main_subject": "1002",
        "main_business_type": "货款",
        "fee_subject": "660305",
        "fee_business_type": "手续费",
    }


def normalize_header_probe_label(label):
    text = str(label or "").strip()
    return HEADER_PROBE_LABEL_KEYS.get(text.lower(), text)


def run_receipt_new_probe():
    return run_receipt_new_probe_with_jab()


def run_receipt_new_probe_with_jab(jab=None):
    if jab is not None:
        from argparse import Namespace
        from tools import receipt_new_probe

        args = Namespace(
            config="config.json",
            method="button",
            path=None,
            title=None,
            class_name="SunAwtFrame",
            name="新增",
            role=None,
            action=None,
            return_timeout=0.2,
            wait=0.8,
            choose_self_made=True,
            self_made_index=0,
            json=False,
            summary=True,
        )
        report = receipt_new_probe.run(args, jab=jab)
        entry_state = report.get("entry_state") or {}
        ok = bool(
            (report.get("open") or {}).get("ok")
            and (report.get("choose_self_made") or {}).get("ok")
        )
        return {
            "step": "open_self_made",
            "ok": ok,
            "parsed": report,
            "entry_state": entry_state,
            "windows_after_choose": report.get("windows_after_choose"),
            "windows_after_open": report.get("windows_after_open"),
            "reason": None
            if ok
            else (
                ((report.get("open") or {}).get("reason"))
                or ((report.get("choose_self_made") or {}).get("reason"))
                or "未能完成新增->自制"
            ),
        }

    cmd = [
        sys.executable,
        str(ROOT / "tools" / "receipt_new_probe.py"),
        "--method",
        "button",
        "--class-name",
        "SunAwtFrame",
        "--choose-self-made",
        "--wait",
        "0.8",
        "--summary",
    ]
    proc = subprocess.run(
        cmd,
        text=True,
        encoding="utf-8",
        errors="replace",
        capture_output=True,
    )
    parsed = None
    ok = False
    try:
        parsed = json.loads(proc.stdout)
        entry_state = parsed.get("entry_state") or {}
        ok = (
            proc.returncode == 0
            and bool((parsed.get("open") or {}).get("ok"))
            and bool((parsed.get("choose_self_made") or {}).get("ok"))
        )
    except json.JSONDecodeError:
        ok = False
    return {
        "step": "open_self_made",
        "ok": ok,
        "returncode": proc.returncode,
        "parsed": parsed,
        "stdout": proc.stdout,
        "stderr": proc.stderr,
    }


def detect_existing_self_made_entry(config):
    jab = JABOperator(config)
    jab.hide_blank_awt_windows_enabled = False
    try:
        jab.ensure_started()
        windows = collect_receipt_new_windows(jab)
        entry_state = detect_self_made_entry_state(windows)
        java_windows = [window for window in windows if window.get("is_java")]
        return {
            "step": "open_self_made",
            "ok": bool(entry_state.get("ok")),
            "method": "existing_entry_state_probe",
            "entry_state": entry_state,
            "java_window_count": len(java_windows),
            "reason": (
                "already in self-made entry state"
                if entry_state.get("ok")
                else "self-made entry state not detected before opening"
            ),
        }
    finally:
        jab.hide_blank_awt_windows_enabled = False
        jab.close()


def probe_header_path_transitions(config):
    jab = JABOperator(config)
    jab.hide_blank_awt_windows_enabled = False
    stages = []
    try:
        jab.ensure_started()
        stages.append(probe_current_header_paths(jab, "before_change"))
        input("已完成第 1 次只读探测。请人工写入/确认财务组织后按回车继续第 2 次探测: ")
        stages.append(probe_current_header_paths(jab, "after_finance"))
        input("已完成第 2 次只读探测。请人工新增/切换窗口后按回车继续第 3 次探测: ")
        stages.append(probe_current_header_paths(jab, "after_window_change"))
        return {
            "ok": any(stage.get("ok") for stage in stages),
            "mode": "header-path-transition-probe",
            "readonly": True,
            "stages": stages,
        }
    finally:
        jab.hide_blank_awt_windows_enabled = False
        jab.close()


def probe_header_semantic_field_speed(config, label, timeout=None, repeat=1):
    jab = JABOperator(config)
    jab.hide_blank_awt_windows_enabled = False
    attempts = []
    timeout = HEADER_LIVE_SEMANTIC_FALLBACK_TIMEOUT if timeout is None else timeout
    repeat = max(int(repeat or 1), 1)
    try:
        jab.ensure_started()
        for index in range(repeat):
            started_at = time.perf_counter()
            found = find_receipt_header_field_by_semantic_label(
                jab,
                label,
                timeout=timeout,
            )
            seconds = round(time.perf_counter() - started_at, 3)
            attempt = {
                "ok": bool(found.get("ok")),
                "index": index + 1,
                "label": label,
                "timeout": timeout,
                "seconds": seconds,
                "source": found.get("source") or "semantic-label-path",
                "path": found.get("path"),
                "label_path": found.get("label_path"),
                "window": found.get("window"),
                "reason": found.get("reason"),
            }
            attempts.append(attempt)
            if found.get("ok"):
                jab.release_contexts(found["vm_id"], found["owned_contexts"])
        successful = [item for item in attempts if item.get("ok")]
        return {
            "ok": bool(successful),
            "mode": "header-semantic-field-speed",
            "readonly": True,
            "label": label,
            "timeout": timeout,
            "repeat": repeat,
            "success_count": len(successful),
            "min_seconds": min((item["seconds"] for item in successful), default=None),
            "max_seconds": max((item["seconds"] for item in successful), default=None),
            "attempts": attempts,
        }
    finally:
        jab.hide_blank_awt_windows_enabled = False
        jab.close()


def probe_customer_name_readback(config, timeout=None):
    jab = JABOperator(config)
    jab.hide_blank_awt_windows_enabled = False
    timeout = HEADER_LIVE_SEMANTIC_FALLBACK_TIMEOUT if timeout is None else timeout
    try:
        jab.ensure_started()
        scope = resolve_current_header_scope_for_probe(jab, timeout=timeout)
        candidates = []
        if scope.get("ok"):
            candidates.extend(
                collect_customer_field_candidates_for_scope(
                    jab,
                    scope.get("scope_hwnd"),
                    scope.get("dynamic_index"),
                    timeout=timeout,
                )
            )
        best = first_valid_customer_name(candidates)
        return {
            "ok": bool(best),
            "mode": "customer-name-readback-probe",
            "readonly": True,
            "scope": scope,
            "best": best,
            "field_candidates": candidates,
            "reason": None if best else _customer_readback_failure_reason(scope, candidates),
        }
    finally:
        jab.hide_blank_awt_windows_enabled = False
        jab.close()


def _customer_readback_failure_reason(scope, candidates):
    """区分客户名回读失败的类别，给人可读原因 + 下一步。"""
    if not scope.get("ok"):
        return (
            "未读到 NC 客户名称：当前页面未定位到收款单表头（可能未停在收款单自制录入页"
            "或表头尚未加载/超时）；请确认 NC 停在收款单自制录入界面后重试。"
        )
    if not candidates:
        return (
            "未读到 NC 客户名称：已定位到表头但客户字段为空；"
            "请确认客户编码已写入并由 NC 带出客户名后再回读。"
        )
    return (
        "未读到 NC 客户名称：读到客户字段内容但格式不符合有效客户名（疑似编码/句柄占位）；"
        "请确认 NC 已用客户编码带出正式客户名后重试。"
    )


def resolve_current_header_scope_for_probe(jab, timeout=None):
    windows = collect_receipt_new_windows(jab)
    state = detect_self_made_entry_state(windows)
    scope_hwnd = None
    for hit in state.get("hits") or []:
        window = hit.get("window") or {}
        if window.get("class_name") == "SunAwtCanvas" and window.get("hwnd"):
            scope_hwnd = int(window["hwnd"])
            break
    if scope_hwnd:
        anchor = resolve_receipt_header_anchor_in_canvas(
            jab,
            scope_hwnd,
            timeout=timeout or 0.6,
        )
        if anchor.get("ok"):
            return anchor
    semantic = infer_receipt_header_scope_by_semantic(jab, scope_hwnd=scope_hwnd)
    if semantic.get("ok"):
        semantic["entry_state"] = state
        return semantic
    return {
        "ok": False,
        "reason": "当前页面未解析到收款单表头 scope",
        "scope_hwnd": scope_hwnd,
        "entry_state": state,
        "semantic_attempt": semantic,
    }


def probe_current_header_paths(jab, stage):
    started_at = time.perf_counter()
    windows = collect_receipt_new_windows(jab)
    candidates = []
    for window in windows:
        if not window.get("visible") or window.get("class_name") != "SunAwtCanvas":
            continue
        hwnd = window.get("hwnd")
        if not hwnd:
            continue
        anchor = resolve_receipt_header_anchor_in_canvas(jab, int(hwnd), timeout=0.15)
        if not anchor.get("ok"):
            continue
        dynamic_index = anchor.get("dynamic_index")
        samples = probe_header_field_paths_for_scope(
            jab,
            int(hwnd),
            dynamic_index,
        )
        candidates.append(
            {
                "scope_hwnd": int(hwnd),
                "dynamic_index": dynamic_index,
                "dynamic_prefix": receipt_header_dynamic_prefix(dynamic_index),
                "anchor": anchor,
                "fields": samples,
            }
        )
    return {
        "ok": bool(candidates),
        "stage": stage,
        "seconds": round(time.perf_counter() - started_at, 3),
        "candidate_count": len(candidates),
        "candidates": candidates,
    }


def probe_header_field_paths_for_scope(jab, scope_hwnd, dynamic_index):
    fields = {}
    for label in HEADER_REQUIRED_LABELS:
        fields[label] = {
            "default_path": summarize_header_path_attempt(
                jab,
                label,
                dynamic_index,
                scope_hwnd,
                template=receipt_header_default_path_template(),
            ),
            "semantic_path": summarize_semantic_header_attempt(
                jab,
                label,
                scope_hwnd,
            ),
        }
    return fields


def summarize_header_path_attempt(jab, label, dynamic_index, scope_hwnd, template):
    if label == HEADER_SCOPE_ANCHOR_LABEL:
        path = build_receipt_header_dynamic_path(dynamic_index, label)
        label_path = build_receipt_header_dynamic_label_path(dynamic_index, label)
        template_source = "finance-anchor"
    else:
        path = build_receipt_header_path_from_template(
            dynamic_index,
            label,
            template,
        )
        label_path = build_receipt_header_label_path_from_template(
            dynamic_index,
            label,
            template,
        )
        template_source = (template or {}).get("source")
    if not path:
        return {
            "ok": False,
            "label": label,
            "reason": "path not configured",
            "template_source": template_source,
        }
    found = find_context_by_path_readonly(
        jab,
        path,
        scope_hwnd=scope_hwnd,
        role="text",
    )
    return {
        "label": label,
        "ok": bool(found.get("ok")),
        "path": path,
        "label_path": label_path,
        "template_source": template_source,
        **found,
    }


def summarize_semantic_header_attempt(jab, label, scope_hwnd):
    found = find_receipt_header_field_by_semantic_label(
        jab,
        label,
        scope_hwnd=scope_hwnd,
    )
    summary = {
        "label": label,
        "ok": bool(found.get("ok")),
        "source": found.get("source") or "semantic-label-path",
        "path": found.get("path"),
        "label_path": found.get("label_path"),
        "window": found.get("window"),
        "reason": found.get("reason"),
    }
    if found.get("ok"):
        inferred = infer_header_path_template_from_field(
            found.get("path"),
            extract_receipt_header_dynamic_index(found.get("path")),
            label,
        )
        summary["inferred_template"] = inferred
        jab.release_contexts(found["vm_id"], found["owned_contexts"])
    return summary


def find_context_by_path_readonly(jab, path, scope_hwnd=None, role=None):
    context, vm_id, owned, window_info = jab.find_context_by_path_once(
        path,
        class_name="SunAwtCanvas",
        scope_hwnd=scope_hwnd,
        role=role,
        require_showing=False,
        require_valid_bounds=False,
    )
    if not context:
        return {"ok": False, "reason": "path not found"}
    try:
        info = jab.get_context_info(vm_id, context)
        text = jab.get_text_context_value(vm_id, context)
        return {
            "ok": True,
            "window": window_info,
            "name": info.name.strip() if info else "",
            "description": info.description.strip() if info else "",
            "role": (info.role_en_US.strip() or info.role.strip()) if info else "",
            "states": (
                (info.states_en_US.strip() or info.states.strip()) if info else ""
            ),
            "text": text,
        }
    finally:
        jab.release_contexts(vm_id, owned)


def fill_detail_line(jab, business):
    steps = []
    for col, name, value in [
        (1, "收款业务类型", business["main_business_type"]),
        (4, "收款银行账户", business["bank_account"]),
        (5, "科目", business["main_subject"]),
        (7, "贷方原币金额", business["amount"]),
        (11, "结算方式", business["settlement"]),
    ]:
        steps.append(fill_cell(jab, 0, col, name, value))
    return steps


def fill_cell(jab, row, col, name, value):
    attempts = []
    for mode, kwargs in [
        (
            "direct",
            {"set_cell_text": value, "commit_key": "none", "focus_target": "cell"},
        ),
    ]:
        try:
            result = select_cell(
                jab,
                table_index=0,
                row=row,
                col=col,
                window_title=None,
                locate_body_table=True,
                wait=0.35,
                **kwargs,
            )
        except Exception as exc:
            result = {"ok": False, "exception": repr(exc)}
        attempts.append({"mode": mode, "result": compact_selection_result(result)})
        if cell_changed(result, value):
            break
        time.sleep(0.2)
    ok = any(cell_changed(attempt["result"], value) for attempt in attempts)
    return {
        "step": "detail_cell",
        "ok": ok,
        "blocked": not ok,
        "reason": (
            None
            if ok
            else "backend cell input failed; global keyboard input is disabled"
        ),
        "row": row,
        "col": col,
        "name": name,
        "value": value,
        "attempts": attempts,
    }


def compact_selection_result(result):
    if not isinstance(result, dict):
        return result
    keep = {
        key: result.get(key)
        for key in (
            "ok",
            "reason",
            "row_count",
            "col_count",
            "child_index",
            "selected_before",
            "selected_after",
            "cell_text_before",
            "cell_text_after",
            "edit",
            "exception",
        )
        if key in result
    }
    return keep


def cell_changed(result, value):
    after = str((result or {}).get("cell_text_after") or "")
    return (
        after
        and after != str((result or {}).get("cell_text_before") or "")
        and str(value) in after
    )


if __name__ == "__main__":
    raise SystemExit(main())
