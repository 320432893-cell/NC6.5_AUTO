from datetime import date
from decimal import Decimal

from openpyxl import Workbook, load_workbook

from core.receipt_config import ReceiptEntryConfig
from core.receipt_entry import ReceiptEntryWorkbook
from core.receipt_matching import (
    ReceiptEntryDryRunMatcher,
    ReceiptEntryMatcher,
    format_receipt_amount_name_mismatch_reason,
    format_receipt_duplicate_reason,
    format_receipt_name_amount_mismatch_reason,
    format_receipt_not_found_reason,
    names_match,
    normalize_counterparty,
)
from core.receipt_models import (
    ReceiptBatchResultRow,
    ReceiptExcelRow,
    ReceiptNCIndexedRow,
    ReceiptNCRow,
)
from core.receipt_nc_extract import ReceiptNCResultExtractor, extract_receipt_nc_rows
from core.receipt_parsing import parse_amount
from core.receipt_parsing import parse_amount as parse_amount_from_new_module
from core.receipt_sheet import RESULT_SHEET_HEADERS


def receipt_config(path="unused.xlsx"):
    return {
        "receipt_entry": {
            "state_label": "收款单录入",
            "excel": {
                "path": str(path),
                "sheet_name": "💸Payments来款通知",
                "header_row": 1,
                "start_row": 2,
                "result_sheet_name": "收款单自动化结果",
                "start_date": "2026-01-01",
                "date_column": "到款日期",
                "payer_name_column": "🟪银行来款名",
                "raw_amount_column": "🟪原始金额",
                "bank_column": "银行",
                "currency_column": "币种",
                "customer_code_column": "客户编码",
                "fee_column": "手续费",
                "organization_column": "主体名称",
            },
            "validation_policy": {
                "mode": "strict",
                "skip_invalid_rows": False,
            },
            "candidate_check": {
                "recent_months": 2,
                "from_date": None,
            },
            "finance_organizations": [
                {
                    "code": "A001",
                    "name": "上海移为通信技术股份有限公司",
                    "short_name": "移为",
                },
                {
                    "code": "A006",
                    "name": "上海移为通信技术（香港）有限公司",
                    "short_name": "移为香港",
                },
            ],
            "accounts": [
                {
                    "organization_code": "A001",
                    "organization_short_name": "移为",
                    "account_label": "PayPal",
                    "account_no": "paypal",
                },
                {
                    "organization_code": "A006",
                    "organization_short_name": "移为香港",
                    "account_label": "香港花旗",
                    "account_no": "1778667904",
                },
            ],
        }
    }


def test_bank_label_maps_to_organization_case_insensitive():
    config = ReceiptEntryConfig(receipt_config())

    organization = config.organization_for_bank("Paypal")

    assert organization is not None
    assert organization.code == "A001"
    assert organization.name == "上海移为通信技术股份有限公司"


def test_extended_account_alias_maps_to_account_and_candidates():
    raw = receipt_config()
    receipt = raw["receipt_entry"]
    receipt["schema_version"] = 2
    receipt["banks"] = [
        {"id": "cmb", "name": "招商银行", "aliases": ["招行"]},
    ]
    receipt["accounts"].append(
        {
            "id": "cmb_a001",
            "enabled": True,
            "organization_code": "A001",
            "organization_short_name": "移为",
            "bank_id": "cmb",
            "account_label": "大陆招行",
            "account_no": "FTE1219165931831",
            "excel_bank_aliases": ["招商", "招行"],
            "nc_candidates_by_currency": {
                "人民币": ["FTE1219165931831RMB"],
                "*": ["FTE1219165931831"],
            },
            "entry_policy": {
                "account_input": "detail_first",
                "success_rule": "non_empty",
            },
        }
    )
    config = ReceiptEntryConfig(raw)

    account = config.account_for_bank("招行")

    assert account is not None
    assert account.id == "cmb_a001"
    assert config.organization_for_bank("招商").code == "A001"
    assert account.nc_candidates("人民币") == [
        "FTE1219165931831RMB",
        "FTE1219165931831",
    ]


def test_ensure_output_columns_and_subjects(tmp_path):
    path = tmp_path / "payments.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "💸Payments来款通知"
    ws.append(["到款日期", "🟪银行来款名", "🟪原始金额", "银行"])
    ws.append([date(2026, 1, 16), "lamine Mohamed", 225.68, "Paypal"])
    ws.append([date(2025, 12, 31), "old", 1, "Paypal"])
    wb.save(path)
    wb.close()

    rows, candidates, issues = ReceiptEntryWorkbook(
        receipt_config(path)
    ).ensure_output_columns_and_subjects(today=date(2026, 1, 20))

    assert issues == []
    assert len(rows) == 1
    assert candidates == rows
    assert rows[0].organization_code == "A001"

    saved = load_workbook(path)
    ws = saved["💸Payments来款通知"]
    headers = [ws.cell(1, column).value for column in range(1, ws.max_column + 1)]
    assert headers == [
        "到款日期",
        "🟪银行来款名",
        "🟪原始金额",
        "银行",
        "主体名称",
    ]
    assert ws.cell(2, 5).value == "上海移为通信技术股份有限公司"
    assert ws.cell(3, 5).value is None
    saved.close()


def test_candidate_rows_use_recent_months_without_status_column(tmp_path):
    path = tmp_path / "payments.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "💸Payments来款通知"
    ws.append(["到款日期", "🟪银行来款名", "🟪原始金额", "银行", "是否NC已做过"])
    ws.append([date(2026, 3, 31), "old recent excluded", 100, "Paypal", None])
    ws.append([date(2026, 4, 2), "already done", 200, "Paypal", "已做过"])
    ws.append([date(2026, 4, 2), "candidate", 300, "Paypal", None])
    wb.save(path)
    wb.close()

    rows, candidates, issues = ReceiptEntryWorkbook(receipt_config(path)).preview_rows(
        today=date(2026, 6, 2)
    )

    assert issues == []
    assert len(rows) == 3
    assert [row.payer_name for row in candidates] == ["already done", "candidate"]


def test_candidate_from_date_overrides_recent_months(tmp_path):
    path = tmp_path / "payments.xlsx"
    config = receipt_config(path)
    config["receipt_entry"]["candidate_check"]["from_date"] = "2026-05-01"
    wb = Workbook()
    ws = wb.active
    ws.title = "💸Payments来款通知"
    ws.append(["到款日期", "🟪银行来款名", "🟪原始金额", "银行"])
    ws.append([date(2026, 4, 30), "old", 100, "Paypal"])
    ws.append([date(2026, 5, 1), "candidate", 200, "Paypal"])
    wb.save(path)
    wb.close()

    rows, candidates, issues = ReceiptEntryWorkbook(config).preview_rows(
        today=date(2026, 6, 2)
    )

    assert issues == []
    assert len(rows) == 2
    assert [row.payer_name for row in candidates] == ["candidate"]


def test_build_local_plan_reports_precise_local_issues(tmp_path):
    path = tmp_path / "payments.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "💸Payments来款通知"
    ws.append(
        [
            "到款日期",
            "🟪银行来款名",
            "🟪原始金额",
            "银行",
            "币种",
            "客户编码",
            "手续费",
        ]
    )
    ws.append([date(2026, 6, 1), "DUP INC", 100, "Paypal", "USD", "YW001", None])
    ws.append([date(2026, 6, 1), "DUP INC", 100, "Paypal", "USD", "YW001", None])
    ws.append([date(2026, 6, 2), "UNKNOWN", 200, "不存在银行", "USD", "YW002", None])
    ws.append([date(2026, 6, 3), "NO CUSTOMER", 300, "Paypal", "USD", "", None])
    wb.save(path)
    wb.close()

    rows, issues, summary = ReceiptEntryWorkbook(
        receipt_config(path)
    ).build_local_plan()

    assert [row.row for row in rows] == [2, 3]
    assert summary["runnable_rows"] == 0
    assert summary["duplicate_rows"] == [2, 3]
    issue_types = [(issue.excel_row, issue.issue_type) for issue in issues]
    assert (2, "DUPLICATE_EXCEL_ROWS") in issue_types
    assert (3, "DUPLICATE_EXCEL_ROWS") in issue_types
    assert (4, "BANK_ACCOUNT_NOT_CONFIGURED") in issue_types
    assert (5, "CUSTOMER_CODE_EMPTY") in issue_types
    bank_issue = next(
        issue for issue in issues if issue.issue_type == "BANK_ACCOUNT_NOT_CONFIGURED"
    )
    assert bank_issue.stage == "配置识别"
    assert bank_issue.field == "银行"
    assert bank_issue.raw_value == "不存在银行"
    assert (
        bank_issue.config_node
        == "receipt_entry.accounts[*].account_label/aliases/excel_bank_aliases"
    )
    assert "未匹配任何账户配置" in bank_issue.message


def test_build_local_plan_writes_machine_sheet2(tmp_path):
    path = tmp_path / "payments.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "💸Payments来款通知"
    ws.append(
        [
            "到款日期",
            "🟪银行来款名",
            "🟪原始金额",
            "银行",
            "币种",
            "客户编码",
            "手续费",
        ]
    )
    ws.append([date(2026, 6, 1), "OK INC", 100, "Paypal", "USD", "YW001", 0])
    wb.save(path)
    wb.close()

    rows, issues, summary = ReceiptEntryWorkbook(receipt_config(path)).build_local_plan(
        write_sheet=True
    )

    assert issues == []
    assert summary["runnable_rows"] == 1
    saved = load_workbook(path)
    ws = saved["收款单自动化结果"]
    headers = [ws.cell(1, column).value for column in range(1, ws.max_column + 1)]
    assert headers[:15] == [
        "原Sheet1行号",
        "执行主体名称",
        "到款日期",
        "🟪银行来款名",
        "客户编码",
        "NC客户名称",
        "🟪原始金额",
        "手续费",
        "🟪到账金额",
        "币种",
        "银行",
        "收款银行账户",
        "本地预检状态",
        "后验核对状态",
        "异常原因",
    ]
    assert ws.cell(2, 1).value == rows[0].row
    assert ws.cell(2, 2).value == "上海移为通信技术股份有限公司"
    assert ws.cell(2, 13).value == "通过"
    assert ws.cell(2, 14).value in (None, "")
    assert ws.cell(2, 15).value in (None, "")
    saved.close()


def test_build_local_plan_rewrites_sheet2_without_duplicate_appends(tmp_path):
    path = tmp_path / "payments.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "💸Payments来款通知"
    ws.append(
        [
            "到款日期",
            "🟪银行来款名",
            "🟪原始金额",
            "银行",
            "币种",
            "客户编码",
            "手续费",
        ]
    )
    ws.append([date(2026, 6, 1), "OK INC", 100, "Paypal", "USD", "YW001", 0])
    wb.save(path)
    wb.close()

    workbook = ReceiptEntryWorkbook(receipt_config(path))
    workbook.build_local_plan(write_sheet=True)
    workbook.build_local_plan(write_sheet=True)

    saved = load_workbook(path)
    ws = saved["收款单自动化结果"]
    columns = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
    assert ws.max_row == 2
    assert ws.cell(2, 1).value == 2
    assert ws.cell(2, columns["本地预检状态"]).value == "通过"
    saved.close()


def test_write_batch_result_sheet_uses_business_columns_and_sorting(tmp_path):
    path = tmp_path / "payments.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "💸Payments来款通知"
    ws.append(
        [
            "到款日期",
            "🟪银行来款名",
            "🟪原始金额",
            "银行",
            "币种",
            "客户编码",
            "手续费",
        ]
    )
    ws.append([date(2026, 6, 2), "LATE", 100, "Paypal", "USD", "YW001", 0])
    ws.append([date(2026, 6, 1), "EARLY", 200, "香港花旗", "USD", "YW002", 3])
    wb.save(path)
    wb.close()

    rows, _issues, _summary = ReceiptEntryWorkbook(
        receipt_config(path)
    ).build_local_plan()
    by_row = {row.row: row for row in rows}
    ReceiptEntryWorkbook(receipt_config(path)).write_batch_result_sheet(
        [
            ReceiptBatchResultRow(
                plan_row=by_row[2],
                local_status="通过",
                nc_customer_name="NC LATE",
                nc_document_no="SK2",
            ),
            ReceiptBatchResultRow(
                plan_row=by_row[3],
                local_status="通过",
                nc_customer_name="NC EARLY",
                nc_document_no="SK1",
            ),
        ]
    )

    saved = load_workbook(path)
    ws = saved["收款单自动化结果"]
    columns = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
    assert ws.cell(2, 1).value == "主体：上海移为通信技术股份有限公司"
    assert ws.cell(2, 1).fill.fgColor.rgb == "00D9EAF7"
    assert ws.cell(3, 1).value == 2
    assert ws.cell(3, 4).value == "LATE"
    assert ws.cell(3, columns["后验核对状态"]).value == "后验通过"
    assert ws.cell(4, 1).value == "主体：上海移为通信技术（香港）有限公司"
    assert ws.cell(5, 1).value == 3
    assert ws.cell(5, 4).value == "EARLY"
    assert ws.cell(5, 6).value == "NC EARLY"
    assert ws.cell(5, 7).value == "203.00"
    assert ws.cell(5, 8).value == "3.00"
    assert ws.cell(5, 9).value == "200.00"
    assert ws.cell(5, columns["后验核对状态"]).value == "后验通过"
    saved.close()


def test_build_local_plan_writes_multiple_global_issues_to_sheet2(tmp_path):
    path = tmp_path / "payments.xlsx"
    config = receipt_config(path)
    config["receipt_entry"]["excel"]["start_row"] = 1
    wb = Workbook()
    ws = wb.active
    ws.title = "💸Payments来款通知"
    ws.append(["到款日期", "🟪银行来款名", "🟪原始金额", "银行"])
    ws.append([date(2026, 6, 1)])
    wb.save(path)
    wb.close()

    _rows, issues, summary = ReceiptEntryWorkbook(config).build_local_plan(
        write_sheet=True
    )

    assert summary["issues"] >= 2
    assert {issue.issue_type for issue in issues} == {"EXCEL_REQUIRED_COLUMN_MISSING"}
    saved = load_workbook(path)
    ws = saved["收款单自动化结果"]
    columns = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
    assert ws.max_row == 3
    assert [ws.cell(row, columns["本地预检状态"]).value for row in range(2, 4)] == [
        "异常"
    ] * 2
    assert all(ws.cell(row, columns["异常原因"]).value for row in range(2, 4))
    saved.close()


def test_counterparty_normalization_ignores_prefix_and_punctuation():
    assert normalize_counterparty("1/AZUGA INC. AZUGA INC - OPERATING") == (
        "AZUGAINCAZUGAINCOPERATING"
    )
    assert names_match("1/AZUGA INC. AZUGA INC - OPERATING", "AZUGA INC")


def test_receipt_matcher_matches_amount_and_name_even_when_dates_differ():
    excel_row = ReceiptExcelRow(
        row=10,
        receipt_date=date(2026, 1, 16),
        payer_name="1/AZUGA INC. AZUGA INC - OPERATING",
        raw_amount=Decimal("68700.00"),
        bank="大陆花旗",
        organization_code="A001",
        organization_name="上海移为通信技术股份有限公司",
        organization_short_name="移为",
        nc_done_status="",
    )
    nc_row = ReceiptNCRow(
        row_index=3,
        document_date=date(2026, 1, 17),
        customer="AZUGA INC",
        original_amount=Decimal("68700.00"),
    )

    matched, issues = ReceiptEntryMatcher().match([excel_row], [nc_row])

    assert matched == {10: nc_row}
    assert issues == []


def test_receipt_matcher_reports_duplicate_as_exception_issue():
    excel_row = ReceiptExcelRow(
        row=10,
        receipt_date=date(2026, 1, 16),
        payer_name="AZUGA INC",
        raw_amount=Decimal("68700.00"),
        bank="大陆花旗",
        organization_code="A001",
        organization_name="上海移为通信技术股份有限公司",
        organization_short_name="移为",
        nc_done_status="",
    )
    nc_rows = [
        ReceiptNCRow(
            row_index=3,
            document_date=date(2026, 1, 16),
            customer="AZUGA INC",
            original_amount=Decimal("68700.00"),
        ),
        ReceiptNCRow(
            row_index=4,
            document_date=date(2026, 1, 16),
            customer="AZUGA INC",
            original_amount=Decimal("68700.00"),
        ),
    ]

    matched, issues = ReceiptEntryMatcher().match([excel_row], nc_rows)

    assert matched == {}
    assert len(issues) == 1
    assert issues[0].reason == format_receipt_duplicate_reason(len(nc_rows))


def test_dry_run_matcher_reports_same_name_different_amount():
    excel_row = ReceiptExcelRow(
        row=10,
        receipt_date=date(2026, 3, 31),
        payer_name="Christoff Pretorius",
        raw_amount=Decimal("100.00"),
        bank="大陆花旗",
        organization_code="A001",
        organization_name="上海移为通信技术股份有限公司",
        organization_short_name="移为",
        nc_done_status="",
    )
    nc_row = ReceiptNCIndexedRow(
        row_index=7,
        table_index=2,
        document_no="D7",
        document_date=date(2026, 4, 1),
        original_amount=Decimal("120.00"),
        name="Christoff Pretorius",
    )

    matched, issues = ReceiptEntryDryRunMatcher().match([excel_row], [nc_row])

    assert matched == {}
    assert len(issues) == 1
    assert issues[0].reason == format_receipt_name_amount_mismatch_reason(
        excel_amount=Decimal("100.00"),
        excel_name="Christoff Pretorius",
        nc_amounts=[Decimal("120.00")],
    )
    assert issues[0].nc_rows == [7]


def test_dry_run_matcher_reports_same_amount_different_name():
    excel_row = ReceiptExcelRow(
        row=10,
        receipt_date=date(2026, 3, 31),
        payer_name="Christoff Pretorius",
        raw_amount=Decimal("100.00"),
        bank="大陆花旗",
        organization_code="A001",
        organization_name="上海移为通信技术股份有限公司",
        organization_short_name="移为",
        nc_done_status="",
    )
    nc_row = ReceiptNCIndexedRow(
        row_index=8,
        table_index=2,
        document_no="D8",
        document_date=date(2026, 4, 1),
        original_amount=Decimal("100.00"),
        name="Different Payer",
    )

    matched, issues = ReceiptEntryDryRunMatcher().match([excel_row], [nc_row])

    assert matched == {}
    assert len(issues) == 1
    assert issues[0].reason == format_receipt_amount_name_mismatch_reason(
        excel_amount=Decimal("100.00"),
        excel_name="Christoff Pretorius",
        nc_names=["Different Payer"],
    )
    assert issues[0].nc_rows == [8]


def test_dry_run_matcher_reports_no_amount_or_name_match():
    excel_row = ReceiptExcelRow(
        row=10,
        receipt_date=date(2026, 3, 31),
        payer_name="Christoff Pretorius",
        raw_amount=Decimal("100.00"),
        bank="大陆花旗",
        organization_code="A001",
        organization_name="上海移为通信技术股份有限公司",
        organization_short_name="移为",
        nc_done_status="",
    )
    nc_row = ReceiptNCIndexedRow(
        row_index=9,
        table_index=2,
        document_no="D9",
        document_date=date(2026, 4, 1),
        original_amount=Decimal("120.00"),
        name="Different Payer",
    )

    matched, issues = ReceiptEntryDryRunMatcher().match([excel_row], [nc_row])

    assert matched == {}
    assert len(issues) == 1
    assert issues[0].reason == format_receipt_not_found_reason()
    assert issues[0].nc_rows == []


def test_extract_receipt_nc_rows_uses_header_labels_not_fixed_indexes():
    tables = [
        {
            "table_index": 2,
            "rows": [
                {
                    "row_index": 0,
                    "cells": ["选择", "客户", "备注", "原币金额", "单据日期"],
                },
                {
                    "row_index": 1,
                    "cells": ["", "AZUGA INC", "", "68,700.00", "2026-01-16"],
                },
                {
                    "row_index": 2,
                    "cells": ["", "", "", "", ""],
                },
            ],
        }
    ]

    rows, issues = extract_receipt_nc_rows(
        tables,
        {
            "document_date": "单据日期",
            "original_amount": "原币金额",
            "customer": "客户",
        },
    )

    assert issues == []
    assert rows == [
        ReceiptNCRow(
            row_index=1,
            document_date=date(2026, 1, 16),
            customer="AZUGA INC",
            original_amount=Decimal("68700.00"),
        )
    ]


def test_extract_receipt_nc_rows_reports_bad_result_values():
    tables = [
        {
            "table_index": 2,
            "rows": [
                {
                    "row_index": 0,
                    "cells": ["单据日期", "客户", "原币金额"],
                },
                {
                    "row_index": 1,
                    "cells": ["2026-01-16", "AZUGA INC", "not money"],
                },
            ],
        }
    ]

    rows, issues = extract_receipt_nc_rows(
        tables,
        {
            "document_date": "单据日期",
            "original_amount": "原币金额",
            "customer": "客户",
        },
    )

    assert rows == []
    assert len(issues) == 1
    assert issues[0].table_index == 2
    assert issues[0].row_index == 1
    assert "原始金额格式无法识别" in issues[0].reason


def test_parse_amount_accepts_nc_negative_spacing():
    assert parse_amount("- 1,368.10") == Decimal("-1368.10")
    assert parse_amount("(1,368.10)") == Decimal("-1368.10")


def test_receipt_split_modules_export_new_paths():
    assert parse_amount_from_new_module("- 1,368.10") == Decimal("-1368.10")
    assert "异常原因" in RESULT_SHEET_HEADERS


def test_extract_receipt_nc_rows_reports_missing_header():
    rows, issues = extract_receipt_nc_rows(
        [{"table_index": 1, "rows": [{"row_index": 0, "cells": ["日期", "金额"]}]}],
        {
            "document_date": "单据日期",
            "original_amount": "原币金额",
            "customer": "客户",
        },
    )

    assert rows == []
    assert len(issues) == 1
    assert issues[0].table_index is None
    assert "未找到包含结果列" in issues[0].reason


def test_receipt_nc_result_extractor_uses_configured_result_column_names():
    config = receipt_config()
    config["receipt_entry"]["query"] = {
        "result_columns": {
            "document_date": "日期",
            "original_amount": "金额",
            "customer": "往来客户",
        }
    }
    tables = [
        {
            "table_index": 3,
            "rows": [
                {"row_index": 0, "cells": ["往来客户", "金额", "日期"]},
                {"row_index": 1, "cells": ["客户A", "100", "2026/01/17"]},
            ],
        }
    ]

    rows, issues = ReceiptNCResultExtractor(config).extract(tables)

    assert issues == []
    assert rows == [
        ReceiptNCRow(
            row_index=1,
            document_date=date(2026, 1, 17),
            customer="客户A",
            original_amount=Decimal("100.00"),
        )
    ]


def test_extract_receipt_nc_rows_by_configured_indexes():
    tables = [
        {
            "table_index": 2,
            "row_count": 1,
            "col_count": 41,
            "rows": [
                {
                    "row_index": 0,
                    "cells": [
                        "D1",
                        "2026-03-31",
                        "收款单",
                        "收款结算",
                        "NC CUSTOMER",
                        "",
                        "37,165.00",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "memo",
                        "BANK PAYER",
                    ],
                }
            ],
        }
    ]
    config = receipt_config()
    config["receipt_entry"]["query"] = {
        "result_column_indexes": {
            "document_no": 0,
            "document_date": 1,
            "customer": 4,
            "original_amount": 6,
            "payer_name": 19,
        }
    }

    rows, issues = ReceiptNCResultExtractor(config).extract_by_indexes(tables, 19)

    assert issues == []
    assert len(rows) == 1
    assert rows[0].document_no == "D1"
    assert rows[0].document_date == date(2026, 3, 31)
    assert rows[0].original_amount == Decimal("37165.00")
    assert rows[0].name == "BANK PAYER"
    assert rows[0].table_index == 2


def test_extract_receipt_nc_rows_by_indexes_collects_all_paged_tables():
    config = receipt_config()
    config["receipt_entry"]["query"] = {
        "result_column_indexes": {
            "document_no": 0,
            "document_date": 1,
            "customer": 2,
            "original_amount": 8,
            "payer_name": 2,
        }
    }
    tables = [
        {
            "table_index": 4,
            "row_count": 2,
            "col_count": 32,
            "rows": [
                {
                    "row_index": 0,
                    "cells": [
                        "D1",
                        "2026-03-31",
                        "PAYER A",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "100.00",
                    ],
                },
                {
                    "row_index": 1,
                    "cells": [
                        "D2",
                        "2026-04-01",
                        "PAYER B",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "200.00",
                    ],
                },
            ],
        },
        {
            "table_index": 4,
            "row_count": 2,
            "col_count": 32,
            "rows": [
                {
                    "row_index": 0,
                    "cells": [
                        "D2",
                        "2026-04-01",
                        "PAYER B",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "200.00",
                    ],
                },
                {
                    "row_index": 1,
                    "cells": [
                        "D3",
                        "2026-04-02",
                        "PAYER C",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "300.00",
                    ],
                },
            ],
        },
    ]

    rows, issues = ReceiptNCResultExtractor(config).extract_by_indexes(tables, 2)

    assert issues == []
    assert [row.document_no for row in rows] == ["D1", "D2", "D3"]
    assert [row.original_amount for row in rows] == [
        Decimal("100.00"),
        Decimal("200.00"),
        Decimal("300.00"),
    ]


def test_dry_run_matcher_compares_chosen_nc_name_field():
    excel_row = ReceiptExcelRow(
        row=10,
        receipt_date=date(2026, 3, 31),
        payer_name="BANK PAYER",
        raw_amount=Decimal("37165.00"),
        bank="大陆花旗",
        organization_code="A001",
        organization_name="上海移为通信技术股份有限公司",
        organization_short_name="移为",
        nc_done_status="",
    )
    nc_row = ReceiptNCResultExtractor(
        {
            "receipt_entry": {
                "query": {
                    "result_column_indexes": {
                        "document_no": 0,
                        "document_date": 1,
                        "customer": 4,
                        "original_amount": 6,
                        "payer_name": 19,
                    }
                },
                "finance_organizations": [],
                "accounts": [],
            }
        }
    ).extract_by_indexes(
        [
            {
                "table_index": 2,
                "row_count": 1,
                "col_count": 41,
                "rows": [
                    {
                        "row_index": 0,
                        "cells": [
                            "D1",
                            "2026-03-31",
                            "",
                            "",
                            "NC CUSTOMER",
                            "",
                            "37,165.00",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "BANK PAYER",
                        ],
                    }
                ],
            }
        ],
        19,
    )[0][0]

    matched, issues = ReceiptEntryDryRunMatcher().match([excel_row], [nc_row])

    assert matched == {10: nc_row}
    assert issues == []


def test_dry_run_matcher_ignores_date_and_matches_amount_name():
    excel_row = ReceiptExcelRow(
        row=10,
        receipt_date=date(2026, 3, 31),
        payer_name="BANK PAYER",
        raw_amount=Decimal("37165.00"),
        bank="大陆花旗",
        organization_code="A001",
        organization_name="上海移为通信技术股份有限公司",
        organization_short_name="移为",
        nc_done_status="",
    )
    nc_row = ReceiptNCIndexedRow(
        row_index=0,
        table_index=2,
        document_no="D1",
        document_date=date(2026, 4, 1),
        original_amount=Decimal("37165.00"),
        name="BANK PAYER",
    )

    matched, issues = ReceiptEntryDryRunMatcher().match([excel_row], [nc_row])

    assert matched == {10: nc_row}
    assert issues == []


def test_dry_run_matcher_reports_duplicate_amount_name():
    excel_row = ReceiptExcelRow(
        row=10,
        receipt_date=date(2026, 3, 31),
        payer_name="BANK PAYER",
        raw_amount=Decimal("37165.00"),
        bank="大陆花旗",
        organization_code="A001",
        organization_name="上海移为通信技术股份有限公司",
        organization_short_name="移为",
        nc_done_status="",
    )
    nc_rows = [
        ReceiptNCIndexedRow(
            row_index=0,
            table_index=2,
            document_no="D1",
            document_date=date(2026, 4, 1),
            original_amount=Decimal("37165.00"),
            name="BANK PAYER",
        ),
        ReceiptNCIndexedRow(
            row_index=1,
            table_index=2,
            document_no="D2",
            document_date=date(2026, 4, 2),
            original_amount=Decimal("37165.00"),
            name="1/BANK PAYER",
        ),
    ]

    matched, issues = ReceiptEntryDryRunMatcher().match([excel_row], nc_rows)

    assert matched == {}
    assert len(issues) == 1
    assert issues[0].reason == format_receipt_duplicate_reason(len(nc_rows))
