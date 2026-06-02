import openpyxl
import re
from decimal import Decimal, InvalidOperation

from core.errors import ExcelLockedError
from core.logger import log
from core.models import ExcelVoucherItem


CONCAT_KEY_RE = re.compile(
    r"^\s*([+-]?(?:\d{1,3}(?:,\d{3})+|\d+)(?:\.\d+)?)(?![\d.,])\s*(.+?)\s*$"
)


class DataHandler:
    def __init__(self, config):
        self.cfg = config
        self.excel_path = config["excel_path"]
        self.sheet_my = config.get("sheet_my", "Sheet1")
        self.has_header = config.get("has_header", True)
        batch_cfg = config.get("jab_batch", {})
        self.jab_key_col = batch_cfg.get("key_col", 1)
        self.jab_amount_out_col = batch_cfg.get("amount_out_col", 1)
        self.jab_partner_out_col = batch_cfg.get("partner_out_col", 2)
        self.jab_result_col = batch_cfg.get("result_col", 3)

    def load_jab_batch_data(
        self, skip_filled=True, skip_any_status=False
    ) -> list[ExcelVoucherItem]:
        """读取 Sheet1 的 A 列拼接索引或 A/B 拆分列，保持 Excel 行顺序。"""
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb[self.sheet_my]

        start_row = 2 if self.has_header else 1
        data = []

        for row in range(start_row, ws.max_row + 1):
            raw_key = ws.cell(row=row, column=self.jab_key_col).value
            raw_amount = ws.cell(row=row, column=self.jab_amount_out_col).value
            raw_partner = ws.cell(row=row, column=self.jab_partner_out_col).value
            result = ws.cell(row=row, column=self.jab_result_col).value

            if (
                self._is_blank(raw_key)
                and self._is_blank(raw_amount)
                and self._is_blank(raw_partner)
            ):
                continue

            if skip_any_status and self._has_cell_value(result):
                continue
            if skip_filled and self._looks_like_voucher(result):
                continue

            try:
                amount, partner, source = self.parse_jab_row(
                    raw_key,
                    raw_amount,
                    raw_partner,
                )
            except ValueError as e:
                log.warning(
                    f"行{row} JAB 索引格式错误: "
                    f"key={raw_key!r}, amount={raw_amount!r}, partner={raw_partner!r}, {e}"
                )
                data.append(
                    ExcelVoucherItem(
                        row=row,
                        raw_key=raw_key,
                        raw_amount=raw_amount,
                        raw_partner=raw_partner,
                        amount=None,
                        partner="",
                        voucher=result,
                        source="",
                        parse_error=str(e),
                    )
                )
                continue

            data.append(
                ExcelVoucherItem(
                    row=row,
                    raw_key=raw_key,
                    raw_amount=raw_amount,
                    raw_partner=raw_partner,
                    amount=amount,
                    partner=partner,
                    voucher=result,
                    source=source,
                    parse_error="",
                )
            )

        for item in data:
            item.validate_for_processing(context="excel_load")

        wb.close()
        log.info(f"加载 JAB 批量数据: {len(data)} 条")
        return data

    def parse_jab_row(
        self,
        raw_key,
        raw_amount,
        raw_partner,
    ):
        errors = []

        try:
            selected_key = self.select_jab_concat_candidate(raw_key, raw_partner)
            amount, partner = self.parse_jab_concat_key(selected_key)
            return amount, partner, "concat"
        except ValueError as e:
            errors.append(f"A列拼接索引: {e}")

        result = self._try_parse_split_row(
            raw_amount,
            raw_partner,
            source="split_ab",
            label="A/B拆分列",
        )
        if isinstance(result, tuple):
            return result
        if result:
            errors.append(result)

        raise ValueError("; ".join(errors))

    def select_jab_concat_candidate(self, raw_key, raw_selector):
        if self._is_blank(raw_key):
            return raw_key

        text = str(raw_key).strip()
        if "/" not in text:
            return text

        candidates = [part.strip() for part in text.split("/")]
        if any(not part for part in candidates):
            raise ValueError("A列多候选存在空项")

        if self._is_blank(raw_selector):
            raise ValueError("A列多候选需要B列填写选择序号")

        selector = self.parse_jab_candidate_selector(raw_selector)
        if selector < 1:
            raise ValueError(f"B列选择序号不是正整数: {raw_selector!r}")
        if selector > len(candidates):
            raise ValueError(
                f"B列选择序号越界: 选择{selector}, 候选{len(candidates)}个"
            )

        return candidates[selector - 1]

    def parse_jab_candidate_selector(self, value):
        text = str(value).strip()
        try:
            selector_decimal = Decimal(text)
        except InvalidOperation as e:
            raise ValueError(f"B列选择序号不是正整数: {value!r}") from e

        if selector_decimal != selector_decimal.to_integral_value():
            raise ValueError(f"B列选择序号不是正整数: {value!r}")
        return int(selector_decimal)

    def _try_parse_split_row(self, raw_amount, raw_partner, source, label):
        has_amount = not self._is_blank(raw_amount)
        has_partner = not self._is_blank(raw_partner)
        if not has_amount and not has_partner:
            return ""
        if not has_amount or not has_partner:
            return f"{label}不完整"

        try:
            amount = self.parse_amount(raw_amount)
        except ValueError as e:
            return f"{label}{e}"

        partner = "".join(str(raw_partner).split())
        if not partner:
            return f"{label}对手方为空"
        return amount, partner, source

    def parse_jab_concat_key(self, value) -> tuple[Decimal, str]:
        if self._is_blank(value):
            raise ValueError("A列拼接索引为空")
        text = str(value).strip()
        match = CONCAT_KEY_RE.match(text)
        if not match:
            raise ValueError("需要以金额开头，后面紧跟对手方名称")

        amount_text, partner = match.groups()
        partner = "".join(partner.split())
        if not partner:
            raise ValueError("对手方名称为空")

        amount = self.parse_amount(amount_text)
        return amount, partner

    def parse_amount(self, value) -> Decimal:
        text = str(value).strip().replace(",", "")
        if not text:
            raise ValueError("金额为空")
        try:
            return Decimal(text).quantize(Decimal("0.01"))
        except (InvalidOperation, ValueError) as e:
            raise ValueError(f"金额格式无法识别: {value!r}") from e

    def split_jab_keys_to_columns(self, limit=None):
        """把“金额+对手方”拼接列拆成独立金额列和对手方列。"""
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb[self.sheet_my]

        start_row = 2 if self.has_header else 1
        end_row = ws.max_row
        if limit:
            end_row = min(end_row, start_row + limit - 1)

        if self.has_header:
            ws.cell(row=1, column=self.jab_amount_out_col, value="金额")
            ws.cell(row=1, column=self.jab_partner_out_col, value="对手方")

        updates = 0
        errors = {}
        for row in range(start_row, end_row + 1):
            raw_key = ws.cell(row=row, column=self.jab_key_col).value
            if raw_key is None or (isinstance(raw_key, str) and raw_key.strip() == ""):
                continue

            try:
                raw_selector = ws.cell(row=row, column=self.jab_partner_out_col).value
                selected_key = self.select_jab_concat_candidate(raw_key, raw_selector)
                amount, partner = self.parse_jab_concat_key(selected_key)
            except ValueError as e:
                errors[row] = str(e)
                log.warning(f"行{row} 拼接索引拆分失败: {raw_key!r}, {e}")
                continue

            ws.cell(row=row, column=self.jab_amount_out_col, value=float(amount))
            ws.cell(row=row, column=self.jab_partner_out_col, value=partner)
            updates += 1
            log.info(f"行{row} 拆分索引: amount={amount} partner={partner}")

        self._save_workbook(wb, "拆分金额/对手方列")
        log.info(f"JAB 拼接索引拆分完成: updates={updates}, errors={len(errors)}")
        return {
            "updates": updates,
            "errors": errors,
            "amount_col": self.jab_amount_out_col,
            "partner_col": self.jab_partner_out_col,
        }

    def save_jab_split_columns(self, items: list[ExcelVoucherItem]):
        updates = {}
        for item in items:
            amount = item.amount
            partner = item.partner
            if (
                not item.parse_error
                and amount is not None
                and partner
                and item.source == "concat"
            ):
                updates[item.row] = (amount, partner)
        if not updates:
            return 0

        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb[self.sheet_my]

        if self.has_header:
            ws.cell(row=1, column=self.jab_amount_out_col, value="金额")
            ws.cell(row=1, column=self.jab_partner_out_col, value="对手方")

        for row, (amount, partner) in updates.items():
            ws.cell(row=row, column=self.jab_amount_out_col, value=float(amount))
            ws.cell(row=row, column=self.jab_partner_out_col, value=partner)
            log.info(f"行{row} 写入拆分列: amount={amount} partner={partner}")

        self._save_workbook(wb, "自动拆分金额/对手方列")
        log.info(f"JAB 自动拆分列写入完成: updates={len(updates)}")
        return len(updates)

    def save_jab_results(self, row_values):
        if not row_values:
            return

        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb[self.sheet_my]

        for row, value in row_values.items():
            ws.cell(row=row, column=self.jab_result_col, value=value)
            log.info(f"行{row} 写入结果: col={self.jab_result_col} value={value}")

        self._save_workbook(wb, "写入凭证状态/凭证号")

    def _save_workbook(self, wb, operation):
        try:
            wb.save(self.excel_path)
        except PermissionError as e:
            raise ExcelLockedError(
                f"Excel 文件无法写入，可能正被 WPS/Excel 打开: "
                f"operation={operation} path={self.excel_path}"
            ) from e
        finally:
            wb.close()

    def _has_cell_value(self, value):
        return value is not None and str(value).strip() != ""

    def _is_blank(self, value):
        return value is None or str(value).strip() == ""

    def _looks_like_voucher(self, value):
        if not self._has_cell_value(value):
            return False
        if isinstance(value, int):
            return value > 0
        text = str(value).strip()
        if isinstance(value, float) and value.is_integer():
            text = str(int(value))
        return text.isdigit() and int(text) > 0
