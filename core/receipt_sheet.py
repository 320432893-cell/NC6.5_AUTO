# 职责：提供收款单 Sheet2 结果表的表头维护、行值生成、排序规则和结果区重写
# 不做什么：不打开/保存 Excel 文件，不查询 NC，不执行 JAB/GUI 动作
# 允许依赖层：收款单数据模型对象和 openpyxl worksheet 风格接口
# 谁不应该 import：底层 JAB/NC workflow 模块不应 import

from core.receipt_amounts import receipt_nc_amount, receipt_net_amount

try:
    from openpyxl.styles import Font, PatternFill
except ImportError:  # pragma: no cover - openpyxl is present in runtime/tests
    Font = None
    PatternFill = None


GROUP_FILL = "D9EAF7"
SUMMARY_FILL = "FFF2CC"
SUCCESS_FILL = "E2F0D9"
ERROR_FILL = "FCE4D6"


RESULT_SHEET_HEADERS = [
    "原Sheet1行号",
    "执行主体名称",
    "到款日期",
    "🟪银行来款名",
    "客户编码",
    "NC客户名称",
    "🟪原始金额",
    "手续费",
    "🟪到账金额",
    "币种",
    "银行",
    "收款银行账户",
    "🔷订单PI匹配",
    "本地预检状态",
    "后验核对状态",
    "异常原因",
]

DEPRECATED_RESULT_SHEET_HEADERS = {
    "执行主体编码",
    "执行主体简称",
    "银行",
    "账户配置ID",
    "异常阶段",
    "异常类型",
    "异常字段",
    "原始值",
    "配置节点",
    "异常说明",
    "处理动作",
    "录入结果",
    "保存结果",
    "后验查询结果",
    "总金额",
    "实收金额",
    "银行来款名",
    "NC单据号",
}


def verification_status(result=None, issue=None, status=""):
    if issue is not None:
        return ""
    if result is None:
        return "待后验" if status == "通过" else ""
    if result.exception_reason:
        return "后验未匹配"
    if result.nc_document_no:
        return "后验通过"
    return "后验待确认" if result.local_status == "通过" else ""


def plan_sheet_row(row, issue, status):
    if row is None:
        base = [""] * 13
    else:
        base = [
            row.row,
            row.organization_name,
            row.receipt_date.isoformat(),
            row.payer_name,
            row.customer_code,
            "",
            str(receipt_nc_amount(row)),
            str(row.fee),
            str(receipt_net_amount(row)),
            row.currency,
            row.bank,
            row.account_no,
            row.extra_text_fields.get("商务领款备忘", ""),
        ]
    issue_reason = format_plan_issue_reason(issue)
    return [
        *base,
        status,
        "",
        issue_reason,
    ]


def batch_result_sheet_row(result):
    row = result.plan_row
    return [
        row.row,
        row.organization_name,
        row.receipt_date.isoformat(),
        row.payer_name,
        row.customer_code,
        result.nc_customer_name,
        str(receipt_nc_amount(row)),
        str(row.fee),
        str(receipt_net_amount(row)),
        row.currency,
        row.bank,
        row.account_no,
        row.extra_text_fields.get("商务领款备忘", ""),
        result.local_status,
        verification_status(result=result),
        result.exception_reason,
    ]


def format_plan_issue_reason(issue):
    if issue is None:
        return ""
    return f"本地预检：{plan_issue_summary(issue)}"


def plan_issue_summary(issue):
    summaries = {
        "EXCEL_REQUIRED_COLUMN_MISSING": "缺少必需列",
        "EXCEL_START_ROW_INVALID": "起始行配置错误",
        "DATE_INVALID": "到款日期格式错误",
        "PAYER_NAME_EMPTY": "银行来款名为空",
        "AMOUNT_ZERO_OR_NEGATIVE": "实收金额必须大于0",
        "AMOUNT_INVALID": "实收金额格式错误",
        "BANK_EMPTY": "银行为空",
        "BANK_ACCOUNT_NOT_CONFIGURED": "银行未配置",
        "BANK_ACCOUNT_DISABLED": "银行账户已禁用",
        "ORG_NOT_CONFIGURED": "执行主体未配置",
        "CURRENCY_EMPTY": "币种为空",
        "CURRENCY_UNSUPPORTED": "币种不支持",
        "CUSTOMER_CODE_EMPTY": "客户编码为空",
        "FEE_NEGATIVE": "手续费不能小于0",
        "FEE_INVALID": "手续费格式错误",
        "DETAIL_ACCOUNT_CANDIDATE_MISSING": "收款银行账户候选缺失",
        "DUPLICATE_EXCEL_ROWS": "本批存在重复行",
    }
    return summaries.get(issue.issue_type) or str(issue.message or "预检失败")


def ensure_result_sheet_headers(ws, header_row):
    for column in range(ws.max_column, 0, -1):
        value = ws.cell(header_row, column).value
        text = str(value or "").strip()
        if text in DEPRECATED_RESULT_SHEET_HEADERS:
            ws.delete_cols(column)

    columns = {}
    for column in range(1, ws.max_column + 1):
        value = ws.cell(header_row, column).value
        text = str(value or "").strip()
        if text:
            columns[text] = column
    next_column = max(columns.values(), default=0) + 1
    for header in RESULT_SHEET_HEADERS:
        if header in columns:
            continue
        ws.cell(row=header_row, column=next_column, value=header)
        columns[header] = next_column
        next_column += 1
    return columns


def append_plan_sheet_row(ws, columns, row_number, values):
    row_by_header = dict(zip(RESULT_SHEET_HEADERS, values, strict=True))
    for header, value in row_by_header.items():
        ws.cell(row=row_number, column=columns[header], value=value)
    return row_number + 1


def append_group_separator_row(ws, columns, row_number, organization_name):
    label = f"主体：{organization_name}"
    max_column = max(columns.values(), default=len(RESULT_SHEET_HEADERS))
    for column in range(1, max_column + 1):
        cell = ws.cell(row=row_number, column=column)
        if column == 1:
            cell.value = label
        apply_group_style(cell)
    return row_number + 1


def append_group_summary_row(ws, columns, row_number, organization_name, results):
    if not results:
        return row_number
    row_count = len(results)
    totals = {
        "🟪原始金额": sum(receipt_nc_amount(result.plan_row) for result in results),
        "手续费": sum(result.plan_row.fee for result in results),
        "🟪到账金额": sum(receipt_net_amount(result.plan_row) for result in results),
    }
    max_column = max(columns.values(), default=len(RESULT_SHEET_HEADERS))
    for column in range(1, max_column + 1):
        cell = ws.cell(row=row_number, column=column)
        apply_summary_style(cell)
    ws.cell(row=row_number, column=columns["原Sheet1行号"], value=f"主体合计：{organization_name}")
    ws.cell(row=row_number, column=columns["执行主体名称"], value=f"{row_count} 条")
    for header, value in totals.items():
        ws.cell(row=row_number, column=columns[header], value=str(value))
    return row_number + 1


def apply_group_style(cell):
    if PatternFill is not None:
        cell.fill = PatternFill("solid", fgColor=GROUP_FILL)
    if Font is not None:
        cell.font = Font(bold=True)


def apply_summary_style(cell):
    if PatternFill is not None:
        cell.fill = PatternFill("solid", fgColor=SUMMARY_FILL)
    if Font is not None:
        cell.font = Font(bold=True)


def apply_result_row_style(ws, columns, row_number, values):
    status = dict(zip(RESULT_SHEET_HEADERS, values, strict=True)).get("后验核对状态")
    fill_color = None
    if status == "后验通过":
        fill_color = SUCCESS_FILL
    elif status in {"后验未匹配", "后验待确认"}:
        fill_color = ERROR_FILL
    if not fill_color or PatternFill is None:
        return
    fill = PatternFill("solid", fgColor=fill_color)
    for column in range(1, max(columns.values(), default=0) + 1):
        ws.cell(row=row_number, column=column).fill = fill


def rewrite_batch_result_sheet(wb, sheet_name, header_row, results):
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb[sheet_name]
    columns = ensure_result_sheet_headers(ws, header_row)
    if ws.max_row > header_row:
        ws.delete_rows(header_row + 1, ws.max_row - header_row)
    append_start_row = header_row + 1
    last_org = None
    group_results = []
    for result in sorted(results, key=batch_result_sort_key):
        org_name = result.plan_row.organization_name
        if org_name != last_org:
            append_start_row = append_group_summary_row(
                ws,
                columns,
                append_start_row,
                last_org,
                group_results,
            )
            group_results = []
            append_start_row = append_group_separator_row(
                ws,
                columns,
                append_start_row,
                org_name,
            )
            last_org = org_name
        values = batch_result_sheet_row(result)
        data_row = append_start_row
        append_start_row = append_plan_sheet_row(
            ws,
            columns,
            append_start_row,
            values,
        )
        apply_result_row_style(ws, columns, data_row, values)
        group_results.append(result)
    append_group_summary_row(ws, columns, append_start_row, last_org, group_results)


def rewrite_plan_sheet(wb, sheet_name, header_row, rows, issues):
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb[sheet_name]
    columns = ensure_result_sheet_headers(ws, header_row)
    if ws.max_row > header_row:
        ws.delete_rows(header_row + 1, ws.max_row - header_row)
    append_start_row = header_row + 1
    issues_by_row = {}
    global_issues = []
    for issue in issues:
        if issue.excel_row is None:
            global_issues.append(issue)
        else:
            issues_by_row.setdefault(issue.excel_row, []).append(issue)
    rows_by_number = {row.row: row for row in rows}
    emitted_rows = set()
    for row in sorted(rows, key=plan_sheet_sort_key):
        emitted_rows.add(row.row)
        row_issues = issues_by_row.get(row.row, [])
        if row_issues:
            for issue in row_issues:
                append_start_row = append_plan_sheet_row(
                    ws,
                    columns,
                    append_start_row,
                    plan_sheet_row(row, issue, "异常"),
                )
        else:
            append_start_row = append_plan_sheet_row(
                ws,
                columns,
                append_start_row,
                plan_sheet_row(row, None, "通过"),
            )
    orphan_issue_items = [
        (row_number, row_issues)
        for row_number, row_issues in issues_by_row.items()
        if row_number not in emitted_rows
    ]
    orphan_issue_items.sort(
        key=lambda item: orphan_issue_sort_key(item, rows_by_number)
    )
    for row_number, row_issues in orphan_issue_items:
        if row_number in emitted_rows:
            continue
        for issue in row_issues:
            append_start_row = append_plan_sheet_row(
                ws,
                columns,
                append_start_row,
                plan_sheet_row(rows_by_number.get(row_number), issue, "异常"),
            )
    for issue in global_issues:
        append_start_row = append_plan_sheet_row(
            ws,
            columns,
            append_start_row,
            plan_sheet_row(None, issue, "异常"),
        )


def plan_sheet_sort_key(row):
    return (
        str(row.organization_code or ""),
        str(row.organization_name or ""),
        row.receipt_date,
        int(row.row or 0),
    )


def batch_result_sort_key(result):
    return plan_sheet_sort_key(result.plan_row)


def orphan_issue_sort_key(item, rows_by_number):
    row_number, _row_issues = item
    row = rows_by_number.get(row_number)
    if row is not None:
        return plan_sheet_sort_key(row)
    return ("ZZZ", "", int(row_number or 0))
