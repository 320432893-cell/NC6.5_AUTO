import os
from pathlib import Path

import openpyxl

from core.errors import ExcelLockedError, WorkflowStateError
from core.receipt_config import (
    ReceiptEntryConfig as ReceiptEntryConfig,
    days_in_month as days_in_month,
    default_account_id as default_account_id,
    normalize_candidate_map as normalize_candidate_map,
    subtract_months as subtract_months,
)
from core.receipt_matching import (
    ReceiptEntryDryRunMatcher as ReceiptEntryDryRunMatcher,
    ReceiptEntryMatcher as ReceiptEntryMatcher,
    format_receipt_amount_name_mismatch_reason as format_receipt_amount_name_mismatch_reason,
    format_receipt_duplicate_reason as format_receipt_duplicate_reason,
    format_receipt_name_amount_mismatch_reason as format_receipt_name_amount_mismatch_reason,
    format_receipt_not_found_reason as format_receipt_not_found_reason,
    names_match as names_match,
    normalize_counterparty as normalize_counterparty,
)
from core.receipt_models import (
    ReceiptExcelRow,
    ReceiptMatchIssue,
    ReceiptNCExtractIssue as ReceiptNCExtractIssue,
    ReceiptNCIndexedRow as ReceiptNCIndexedRow,
    ReceiptNCRow as ReceiptNCRow,
    ReceiptPlanIssue as ReceiptPlanIssue,
)
from core.receipt_nc_extract import (
    ReceiptNCResultExtractor as ReceiptNCResultExtractor,
    extract_receipt_nc_rows as extract_receipt_nc_rows,
    extract_receipt_nc_rows_by_indexes as extract_receipt_nc_rows_by_indexes,
    is_blank_result_row as is_blank_result_row,
    read_cell as read_cell,
    read_required_text as read_required_text,
    resolve_receipt_result_columns as resolve_receipt_result_columns,
)
from core.receipt_parsing import (
    make_receipt_duplicate_key as build_receipt_duplicate_key,
    normalize_lookup_key,
    parse_amount,
    parse_date,
)
from core.receipt_plan import (
    build_plan_rows,
)
from core.receipt_plan_issue import (
    detect_duplicate_rows,
    plan_issue as plan_issue,
    read_optional_cell,
    summarize_plan,
)

# 兼容别名：旧调用方仍可从 core.receipt_entry import；待生产脚本和测试迁到
# core.receipt_models/core.receipt_matching/core.receipt_nc_extract/
# core.receipt_parsing/core.receipt_sheet 后删除这些 re-export。
from core.receipt_sheet import (
    DEPRECATED_RESULT_SHEET_HEADERS as DEPRECATED_RESULT_SHEET_HEADERS,
    RESULT_SHEET_HEADERS as RESULT_SHEET_HEADERS,
    append_plan_sheet_row as append_plan_sheet_row,
    ensure_result_sheet_headers as ensure_result_sheet_headers,
    orphan_issue_sort_key as orphan_issue_sort_key,
    plan_sheet_row as plan_sheet_row,
    plan_sheet_sort_key as plan_sheet_sort_key,
    rewrite_batch_result_sheet,
    rewrite_plan_sheet,
)


class ReceiptEntryWorkbook:
    def __init__(self, config, excel_path=None):
        self.config = ReceiptEntryConfig(config)
        workbook_path = excel_path or self.config.excel_cfg.get("path")
        if not workbook_path:
            raise WorkflowStateError("receipt_entry.excel.path is required")
        self.excel_path = str(self._resolve_workbook_path(workbook_path))

    @staticmethod
    def _resolve_workbook_path(workbook_path):
        path = Path(str(workbook_path))
        if path.exists():
            return path
        if path.is_absolute():
            return path

        name = path.name
        candidates = [Path.cwd() / path]
        downloads_roots = []
        home = Path.home()
        downloads_roots.append(home / "Downloads")
        userprofile = os.environ.get("USERPROFILE")
        if userprofile:
            downloads_roots.append(Path(userprofile) / "Downloads")
        if os.environ.get("HOME", "").startswith("/home/"):
            username = Path(os.environ["HOME"]).name
            downloads_roots.append(Path("/mnt/c/Users") / username / "Downloads")
        for root in downloads_roots:
            candidates.append(root / name)
        for candidate in candidates:
            if candidate.exists():
                return candidate
        return path

    def preview_rows(self, today=None):
        wb = openpyxl.load_workbook(self.excel_path, read_only=False, data_only=True)
        try:
            ws = wb[self.config.sheet_name]
            columns = self._read_header(ws)
            rows, issues = self._load_rows(ws, columns)
            return rows, self.select_candidate_rows(rows, today=today), issues
        finally:
            wb.close()

    def build_local_plan(self, write_sheet=False):
        wb = openpyxl.load_workbook(self.excel_path, read_only=False, data_only=True)
        try:
            ws = wb[self.config.sheet_name]
            columns = self._read_header(ws)
            rows, issues = build_plan_rows(self.config, ws, columns)
            duplicate_issues = detect_duplicate_rows(rows)
            issues.extend(duplicate_issues)
        finally:
            wb.close()
        if write_sheet:
            writable = openpyxl.load_workbook(self.excel_path, read_only=False)
            try:
                self._write_plan_sheet(writable, rows, issues)
            finally:
                writable.close()
        return rows, issues, summarize_plan(rows, issues, self.config.validation_policy)

    def write_plan_sheet(self, rows, issues):
        writable = openpyxl.load_workbook(self.excel_path, read_only=False)
        try:
            self._write_plan_sheet(writable, rows, issues)
        finally:
            writable.close()

    def write_batch_result_sheet(self, results):
        writable = openpyxl.load_workbook(self.excel_path, read_only=False)
        try:
            rewrite_batch_result_sheet(
                writable,
                self.config.result_sheet_name,
                self.config.header_row,
                results,
            )
            try:
                writable.save(self.excel_path)
            except PermissionError as exc:
                raise ExcelLockedError(
                    f"Excel 文件无法写入，可能正被 WPS/Excel 打开: "
                    f"operation=写入收款单执行结果 path={self.excel_path}"
                ) from exc
        finally:
            writable.close()

    def ensure_output_columns_and_subjects(self, today=None):
        wb = openpyxl.load_workbook(self.excel_path, read_only=False)
        try:
            ws = wb[self.config.sheet_name]
            columns = self._read_header(ws)
            columns = self._ensure_column(ws, columns, self.config.organization_column)
            rows, issues = self._load_rows(ws, columns)
            org_col = columns[self.config.organization_column]
            for row in rows:
                ws.cell(row=row.row, column=org_col, value=row.organization_name)
            self._save_workbook(wb, "写入收款单主体预处理列")
            return rows, self.select_candidate_rows(rows, today=today), issues
        except PermissionError as exc:
            wb.close()
            raise ExcelLockedError(
                f"Excel 文件无法写入，可能正被 WPS/Excel 打开: path={self.excel_path}"
            ) from exc

    def select_candidate_rows(self, rows, today=None):
        candidate_start = self.config.candidate_start_date(today=today)
        candidates = []
        for row in rows:
            if row.receipt_date < candidate_start:
                continue
            candidates.append(row)
        return candidates

    def _load_rows(self, ws, columns):
        required = [
            self.config.date_column,
            self.config.payer_name_column,
            self.config.raw_amount_column,
            self.config.bank_column,
        ]
        missing = [name for name in required if name not in columns]
        if missing:
            raise WorkflowStateError(f"收款 Excel 缺少必需列: {missing}")

        rows = []
        issues = []
        for row_index in range(self.config.header_row + 1, ws.max_row + 1):
            raw_date = ws.cell(row_index, columns[self.config.date_column]).value
            if raw_date in (None, ""):
                continue
            try:
                receipt_date = parse_date(raw_date)
            except ValueError as exc:
                issues.append(ReceiptMatchIssue(row_index, str(exc), []))
                continue
            if receipt_date < self.config.start_date:
                continue

            bank = ws.cell(row_index, columns[self.config.bank_column]).value
            payer_name = ws.cell(
                row_index, columns[self.config.payer_name_column]
            ).value
            raw_amount = ws.cell(
                row_index, columns[self.config.raw_amount_column]
            ).value
            organization = self.config.organization_for_bank(bank)
            if not organization:
                issues.append(ReceiptMatchIssue(row_index, f"银行未配置: {bank!r}", []))
                continue
            try:
                amount = parse_amount(raw_amount)
            except ValueError as exc:
                issues.append(ReceiptMatchIssue(row_index, str(exc), []))
                continue
            payer_text = "" if payer_name is None else str(payer_name).strip()
            if not payer_text:
                issues.append(ReceiptMatchIssue(row_index, "银行来款名为空", []))
                continue
            rows.append(
                ReceiptExcelRow(
                    row=row_index,
                    receipt_date=receipt_date,
                    payer_name=payer_text,
                    raw_amount=amount,
                    bank="" if bank is None else str(bank).strip(),
                    organization_code=organization.code,
                    organization_name=organization.name,
                    organization_short_name=organization.short_name,
                    nc_done_status="",
                )
            )
        return rows, issues

    def _write_plan_sheet(self, wb, rows, issues):
        rewrite_plan_sheet(
            wb,
            self.config.result_sheet_name,
            self.config.header_row,
            rows,
            issues,
        )
        try:
            wb.save(self.excel_path)
        except PermissionError as exc:
            raise ExcelLockedError(
                f"Excel 文件无法写入，可能正被 WPS/Excel 打开: "
                f"operation=写入收款单计划结果 path={self.excel_path}"
            ) from exc

    def _read_header(self, ws):
        columns = {}
        for column in range(1, ws.max_column + 1):
            value = ws.cell(self.config.header_row, column).value
            if value is None:
                continue
            text = str(value).strip()
            if text:
                columns[text] = column
        return columns

    def _ensure_column(self, ws, columns, name):
        if name in columns:
            return columns
        column = ws.max_column + 1
        ws.cell(row=self.config.header_row, column=column, value=name)
        return {**columns, name: column}

    def _save_workbook(self, wb, operation):
        try:
            wb.save(self.excel_path)
        except PermissionError as exc:
            raise ExcelLockedError(
                f"Excel 文件无法写入，可能正被 WPS/Excel 打开: "
                f"operation={operation} path={self.excel_path}"
            ) from exc
        finally:
            wb.close()


def make_receipt_duplicate_key(
    organization_code,
    receipt_date,
    bank,
    currency,
    customer_code,
    payer_name,
    amount,
):
    # 兼容包装：旧调用方仍传 receipt_entry 的规范化函数；迁到
    # core.receipt_parsing.make_receipt_duplicate_key 后删除。
    return build_receipt_duplicate_key(
        organization_code,
        receipt_date,
        bank,
        currency,
        customer_code,
        payer_name,
        amount,
        normalize_lookup_key,
        normalize_counterparty,
    )
